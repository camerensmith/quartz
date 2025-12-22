"""CSV import dialog with column mapping"""

from pathlib import Path
from typing import List, Dict, Optional
import csv

from PySide6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTableWidget,
    QTableWidgetItem, QComboBox, QFileDialog, QMessageBox, QGroupBox, QCheckBox, QHeaderView
)
from PySide6.QtCore import Qt

# pandas removed - using csv module and openpyxl directly


def detect_file_encoding(file_path: Path) -> str:
    """
    Detect file encoding with fallback options.
    
    Args:
        file_path: Path to the file
        
    Returns:
        Detected encoding string
    """
    # Common encodings to try (in order of likelihood)
    encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1', 'windows-1252']
    
    # Try to detect encoding using chardet if available
    try:
        import chardet
        with open(file_path, 'rb') as f:
            raw_data = f.read(10000)  # Read first 10KB for detection
            if raw_data:
                detected = chardet.detect(raw_data)
                if detected and detected.get('encoding'):
                    detected_encoding = detected['encoding'].lower()
                    # Normalize some common variations
                    if detected_encoding in ['iso-8859-1', 'latin-1']:
                        detected_encoding = 'latin-1'
                    elif detected_encoding in ['windows-1252', 'cp1252']:
                        detected_encoding = 'cp1252'
                    # Add detected encoding to the front of the list
                    if detected_encoding not in encodings:
                        encodings.insert(0, detected_encoding)
                    else:
                        # Move to front
                        encodings.remove(detected_encoding)
                        encodings.insert(0, detected_encoding)
    except (ImportError, Exception):
        pass  # Fall back to default list (chardet not available or detection failed)
    
    # Try each encoding
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                f.read(1024)  # Try reading a small chunk
            return encoding
        except (UnicodeDecodeError, LookupError):
            continue
    
    # Last resort: return utf-8 with errors='replace'
    return 'utf-8'


class ImportDialog(QDialog):
    """Dialog for importing CSV files"""
    
    def __init__(self, parent=None, store=None, fields: Optional[List[Dict]] = None):
        super().__init__(parent)
        self.store = store
        self.fields = fields or []
        self.csv_data = []
        self.csv_headers = []
        self.file_path: Optional[Path] = None
        
        self.setWindowTitle("Import CSV")
        self.setMinimumWidth(700)
        self.setMinimumHeight(500)
        
        # Apply theme from parent if available
        if parent:
            self.setStyleSheet(parent.styleSheet())
        
        self._init_ui()
    
    def _init_ui(self):
        """Initialize UI"""
        layout = QVBoxLayout(self)
        
        # File selection
        file_layout = QHBoxLayout()
        file_layout.addWidget(QLabel("CSV File:"))
        self.file_label = QLabel("(not selected)")
        self.file_label.setStyleSheet("color: gray;")
        file_layout.addWidget(self.file_label)
        file_layout.addStretch()
        
        browse_btn = QPushButton("Browse...")
        browse_btn.clicked.connect(self._browse_file)
        file_layout.addWidget(browse_btn)
        layout.addLayout(file_layout)
        
        # Update label to reflect supported formats
        self.file_label.setText("(CSV or Excel file)")
        
        # Options
        options_group = QGroupBox("Import Options")
        options_layout = QVBoxLayout()
        
        self.skip_first_row_check = QCheckBox("Skip first row (header)")
        self.skip_first_row_check.setChecked(True)
        options_layout.addWidget(self.skip_first_row_check)
        
        self.create_fields_check = QCheckBox("Create new fields for unmapped columns")
        self.create_fields_check.setChecked(False)
        self.create_fields_check.toggled.connect(self._on_create_fields_toggled)
        options_layout.addWidget(self.create_fields_check)
        
        # Button to create all fields from CSV
        create_all_btn = QPushButton("Create All Fields from CSV")
        create_all_btn.clicked.connect(self._create_all_fields_from_csv)
        options_layout.addWidget(create_all_btn)
        
        options_group.setLayout(options_layout)
        layout.addWidget(options_group)
        
        # Column mapping table
        layout.addWidget(QLabel("Column Mapping:"))
        self.mapping_table = QTableWidget()
        self.mapping_table.setColumnCount(4)  # Added Field Type column
        self.mapping_table.setHorizontalHeaderLabels(["CSV Column", "Map to Field", "Field Type", "Sample Data"])
        self.mapping_table.horizontalHeader().setStretchLastSection(True)
        self.mapping_table.setSelectionBehavior(QTableWidget.SelectRows)
        layout.addWidget(self.mapping_table)
        
        # Buttons
        button_layout = QHBoxLayout()
        preview_btn = QPushButton("Preview")
        preview_btn.clicked.connect(self._preview_import)
        button_layout.addWidget(preview_btn)
        button_layout.addStretch()
        
        cancel_btn = QPushButton("Cancel")
        cancel_btn.setProperty("class", "secondary")
        cancel_btn.clicked.connect(self.reject)
        
        import_btn = QPushButton("Import")
        import_btn.setDefault(True)
        import_btn.clicked.connect(self._import_data)
        
        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(import_btn)
        layout.addLayout(button_layout)
    
    def _browse_file(self):
        """Browse for CSV or Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Import CSV or Excel", "", 
            "CSV files (*.csv);;Excel files (*.xlsx *.xls);;All files (*)"
        )
        if file_path:
            self.file_path = Path(file_path)
            self.file_label.setText(self.file_path.name)
            self.file_label.setStyleSheet("")
            self._load_csv()
    
    def _load_csv(self):
        """Load and parse CSV or Excel file"""
        if not self.file_path or not self.file_path.exists():
            return
        
        try:
            file_ext = self.file_path.suffix.lower()
            if file_ext in ('.xlsx', '.xls'):
                # Load Excel file using openpyxl
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(self.file_path, read_only=True, data_only=True)
                    ws = wb.active
                    
                    # Get headers from first row
                    self.csv_headers = []
                    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                    for cell_value in first_row:
                        self.csv_headers.append(str(cell_value) if cell_value is not None else "")
                    
                    # Get data (limit to 100 rows for preview)
                    self.csv_data = []
                    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
                        if i > 100:  # Limit preview
                            break
                        row_data = [str(cell) if cell is not None else "" for cell in row]
                        self.csv_data.append(row_data)
                    
                    wb.close()
                except ImportError:
                    QMessageBox.critical(self, "Error", "openpyxl is required for Excel files. Install it with: pip install openpyxl")
                    return
            else:
                # Load CSV file using built-in csv module
                # Detect encoding first
                encoding = detect_file_encoding(self.file_path)
                try:
                    with open(self.file_path, 'r', encoding=encoding) as f:
                        reader = csv.reader(f)
                        self.csv_headers = next(reader)
                        self.csv_data = []
                        for i, row in enumerate(reader):
                            if i >= 100:  # Limit preview
                                break
                            self.csv_data.append(row)
                except UnicodeDecodeError:
                    # If detected encoding fails, try with error handling
                    with open(self.file_path, 'r', encoding=encoding, errors='replace') as f:
                        reader = csv.reader(f)
                        self.csv_headers = next(reader)
                        self.csv_data = []
                        for i, row in enumerate(reader):
                            if i >= 100:  # Limit preview
                                break
                            self.csv_data.append(row)
            
            self._populate_mapping_table()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to load file:\n{str(e)}")
    
    def _populate_mapping_table(self):
        """Populate column mapping table"""
        self.mapping_table.setRowCount(len(self.csv_headers))
        
        # Get field names
        field_names = ["(ignore)"] + [f["label"] for f in self.fields]
        
        for row, header in enumerate(self.csv_headers):
            # CSV Column
            header_item = QTableWidgetItem(header)
            header_item.setFlags(header_item.flags() & ~Qt.ItemIsEditable)
            self.mapping_table.setItem(row, 0, header_item)
            
            # Map to Field (combo)
            combo = QComboBox()
            combo.addItems(field_names)
            
            # Try to auto-match by name
            matching_field = next(
                (f for f in self.fields
                 if f["key"].lower() == header.lower() or f["label"].lower() == header.lower()),
                None
            )
            if matching_field:
                combo.setCurrentText(matching_field["label"])
            
            self.mapping_table.setCellWidget(row, 1, combo)
            
            # Field Type (for new fields) - only visible if create_fields_check is checked
            field_types = ["text", "notes", "integer", "decimal", "checkbox", "date", "datetime", "dropdown"]
            type_combo = QComboBox()
            type_combo.addItems(field_types)
            # Try to infer type from sample data
            if self.csv_data and len(self.csv_data) > 0 and row < len(self.csv_data[0]):
                sample = str(self.csv_data[0][row]).strip()
                if sample:
                    # Simple type inference
                    sample_lower = sample.lower()
                    if sample_lower in ("true", "false", "1", "0", "yes", "no"):
                        type_combo.setCurrentText("checkbox")
                    elif sample.replace(".", "").replace("-", "").isdigit():
                        if "." in sample:
                            type_combo.setCurrentText("decimal")
                        else:
                            type_combo.setCurrentText("integer")
                    elif len(sample) > 100:
                        type_combo.setCurrentText("notes")
            type_combo.setEnabled(self.create_fields_check.isChecked())
            self.mapping_table.setCellWidget(row, 2, type_combo)
            
            # Sample data
            sample = ""
            if self.csv_data and len(self.csv_data) > 0 and row < len(self.csv_data[0]):
                sample = str(self.csv_data[0][row])[:50]
            
            sample_item = QTableWidgetItem(sample)
            sample_item.setFlags(sample_item.flags() & ~Qt.ItemIsEditable)
            self.mapping_table.setItem(row, 3, sample_item)
        
        # Update column visibility
        self._on_create_fields_toggled(self.create_fields_check.isChecked())
        self.mapping_table.resizeColumnsToContents()
    
    def _on_create_fields_toggled(self, checked: bool):
        """Handle create fields checkbox toggle"""
        # Show/hide field type column and enable/disable type combos
        self.mapping_table.setColumnHidden(2, not checked)
        for row in range(self.mapping_table.rowCount()):
            type_combo = self.mapping_table.cellWidget(row, 2)
            if type_combo:
                type_combo.setEnabled(checked)
    
    def _create_all_fields_from_csv(self):
        """Create fields for all CSV columns that aren't mapped"""
        if not self.csv_headers:
            QMessageBox.warning(self, "Error", "Please load a CSV file first")
            return
        
        # Enable create fields checkbox
        self.create_fields_check.setChecked(True)
        
        # Set all unmapped columns to create new fields
        for row in range(self.mapping_table.rowCount()):
            combo = self.mapping_table.cellWidget(row, 1)
            if combo:
                current_text = combo.currentText()
                # If set to "(ignore)" or not matching an existing field, set to create new
                if current_text == "(ignore)":
                    # Keep as ignore but will be converted to new field on import
                    pass
                else:
                    # Check if it matches an existing field
                    matching_field = next(
                        (f for f in self.fields if f["label"] == current_text),
                        None
                    )
                    if not matching_field:
                        # Not a valid field, set to ignore (will create new field)
                        combo.setCurrentText("(ignore)")
        
        QMessageBox.information(
            self, "Fields Created",
            f"All unmapped CSV columns will be created as new fields.\n"
            f"You can adjust field types in the 'Field Type' column before importing."
        )
    
    def _preview_import(self):
        """Preview import (show how many records will be imported)"""
        if not self.file_path:
            QMessageBox.warning(self, "Error", "Please select a CSV file first")
            return
        
        try:
            # Count total rows
            encoding = detect_file_encoding(self.file_path)
            try:
                with open(self.file_path, 'r', encoding=encoding) as f:
                    reader = csv.reader(f)
                    row_count = sum(1 for _ in reader) - (1 if self.skip_first_row_check.isChecked() else 0)
            except UnicodeDecodeError:
                with open(self.file_path, 'r', encoding=encoding, errors='replace') as f:
                    reader = csv.reader(f)
                    row_count = sum(1 for _ in reader) - (1 if self.skip_first_row_check.isChecked() else 0)
            
            QMessageBox.information(
                self, "Import Preview",
                f"Will import approximately {row_count} records.\n\n"
                f"CSV has {len(self.csv_headers)} columns.\n"
                f"Mapping to {len(self.fields)} existing fields."
            )
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to preview:\n{str(e)}")
    
    def _import_data(self):
        """Perform the import"""
        if not self.file_path or not self.store:
            QMessageBox.warning(self, "Error", "Please select a CSV file and ensure a collection is open")
            return
        
        # Build mapping
        mapping = {}
        new_fields_to_create = []
        
        for row in range(self.mapping_table.rowCount()):
            csv_col = self.mapping_table.item(row, 0).text()
            combo = self.mapping_table.cellWidget(row, 1)
            field_name = combo.currentText()
            
            # Get field type from type combo (column 2)
            type_combo = self.mapping_table.cellWidget(row, 2)
            field_type = type_combo.currentText() if type_combo else "text"
            
            if field_name == "(ignore)":
                # If "create new fields" is checked, still create fields for ignored columns
                if self.create_fields_check.isChecked():
                    # Create new field for this column
                    field_key = csv_col.lower().replace(" ", "_").replace("-", "_")
                    field_key = "".join(c for c in field_key if c.isalnum() or c == "_")
                    if not field_key or field_key[0].isdigit():
                        field_key = f"field_{field_key}" if field_key else f"field_{len(new_fields_to_create)}"
                    new_fields_to_create.append({
                        "key": field_key,
                        "label": csv_col,
                        "type": field_type
                    })
                    mapping[csv_col] = field_key
                # Otherwise, skip this column
                continue
            
            # Find field
            field = next((f for f in self.fields if f["label"] == field_name), None)
            if field:
                mapping[csv_col] = field["key"]
            elif self.create_fields_check.isChecked():
                # Create new field
                field_key = csv_col.lower().replace(" ", "_").replace("-", "_")
                field_key = "".join(c for c in field_key if c.isalnum() or c == "_")
                if not field_key or field_key[0].isdigit():
                    field_key = f"field_{field_key}" if field_key else f"field_{len(new_fields_to_create)}"
                
                # Generate unique key - check against existing fields and already-created new fields
                existing_fields = self.fields.copy()
                existing_fields.extend([{"key": f["key"]} for f in new_fields_to_create])
                from src.ui.add_field_dialog import _generate_unique_field_key
                field_key = _generate_unique_field_key(field_key, existing_fields)
                
                new_fields_to_create.append({
                    "key": field_key,
                    "label": csv_col,
                    "type": field_type
                })
                mapping[csv_col] = field_key
        
        # Validation: need at least one mapped column
        # If "create new fields" is checked, columns set to "(ignore)" should have been converted to new fields
        if not mapping:
            if self.create_fields_check.isChecked():
                # This shouldn't happen if create_fields is checked and CSV has columns
                QMessageBox.warning(self, "Error", "No columns to import. Please ensure your CSV file has columns.")
            else:
                QMessageBox.warning(self, "Error", "Please map at least one column to a field, or enable 'Create new fields for unmapped columns'")
            return
        
        # Create new fields if needed
        for field_def in new_fields_to_create:
            try:
                self.store.add_field(
                    field_key=field_def["key"],
                    field_type=field_def["type"],
                    label=field_def["label"],
                    required=False
                )
            except Exception as e:
                QMessageBox.warning(self, "Warning", f"Failed to create field {field_def['label']}:\n{str(e)}")
        
        # Import records
        try:
            imported = 0
            errors = []
            
            file_ext = self.file_path.suffix.lower()
            if file_ext in ('.xlsx', '.xls'):
                # Import from Excel using openpyxl
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(self.file_path, read_only=True, data_only=True)
                    ws = wb.active
                    
                    # Determine start row based on skip_first_row_check
                    start_row = 2 if self.skip_first_row_check.isChecked() else 1
                    
                    # Get headers
                    if self.skip_first_row_check.isChecked():
                        # Headers from first row
                        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                        excel_headers = [str(cell) if cell is not None else f"Column_{i+1}" for i, cell in enumerate(header_row)]
                    else:
                        # Generate column names
                        first_data_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                        excel_headers = [f"Column_{i+1}" for i in range(len(first_data_row))]
                        start_row = 1
                    
                    # Import rows
                    for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
                        try:
                            # Convert row to dict
                            row_dict = {}
                            for i, cell_value in enumerate(row):
                                if i < len(excel_headers):
                                    col_name = excel_headers[i]
                                    row_dict[col_name] = str(cell_value) if cell_value is not None else ""
                            
                            # Map to field keys
                            record_data = {}
                            for csv_col, field_key in mapping.items():
                                value = row_dict.get(csv_col, "")
                                if value and value.strip():  # Check for empty/None
                                    record_data[field_key] = value
                            
                            # Add record
                            if record_data:
                                self.store.add_record(record_data)
                                imported += 1
                        except Exception as e:
                            errors.append(f"Row {row_idx}: {str(e)}")
                            if len(errors) >= 10:  # Limit error messages
                                break
                    
                    wb.close()
                except ImportError:
                    QMessageBox.critical(self, "Error", "openpyxl is required for Excel files. Install it with: pip install openpyxl")
                    return
            else:
                # Import from CSV
                encoding = detect_file_encoding(self.file_path)
                # Use headers from preview (self.csv_headers) - these are the actual CSV column names
                csv_headers = self.csv_headers if self.csv_headers else []
                
                if not csv_headers:
                    QMessageBox.warning(self, "Error", "No CSV headers found. Please reload the file.")
                    return
                
                try:
                    with open(self.file_path, 'r', encoding=encoding) as f:
                        reader = csv.reader(f)
                        # Skip first row if header checkbox is checked
                        if self.skip_first_row_check.isChecked():
                            next(reader, None)  # Skip header row
                        
                        # Now iterate over all data rows (while file is still open)
                        for row_idx, row in enumerate(reader, start=1):
                            try:
                                # Convert row list to dict using headers
                                row_dict = {}
                                for i, val in enumerate(row):
                                    if i < len(csv_headers):
                                        row_dict[csv_headers[i]] = str(val) if val is not None else ""
                                
                                # Map to field keys
                                record_data = {}
                                for csv_col, field_key in mapping.items():
                                    value = row_dict.get(csv_col, "")
                                    if value and str(value).strip():  # Only add non-empty values
                                        record_data[field_key] = str(value).strip()
                                
                                # Add record if we have any data
                                if record_data:
                                    self.store.add_record(record_data)
                                    imported += 1
                            except Exception as e:
                                errors.append(f"Row {row_idx+1}: {str(e)}")
                                if len(errors) >= 10:  # Limit error messages
                                    break
                except UnicodeDecodeError:
                    with open(self.file_path, 'r', encoding=encoding, errors='replace') as f:
                        reader = csv.reader(f)
                        if self.skip_first_row_check.isChecked():
                            next(reader, None)  # Skip header row
                        
                        # Now iterate over all data rows (while file is still open)
                        for row_idx, row in enumerate(reader, start=1):
                            try:
                                # Convert row list to dict using headers
                                row_dict = {}
                                for i, val in enumerate(row):
                                    if i < len(csv_headers):
                                        row_dict[csv_headers[i]] = str(val) if val is not None else ""
                                
                                # Map to field keys
                                record_data = {}
                                for csv_col, field_key in mapping.items():
                                    value = row_dict.get(csv_col, "")
                                    if value and str(value).strip():  # Only add non-empty values
                                        record_data[field_key] = str(value).strip()
                                
                                # Add record if we have any data
                                if record_data:
                                    self.store.add_record(record_data)
                                    imported += 1
                            except Exception as e:
                                errors.append(f"Row {row_idx+1}: {str(e)}")
                                if len(errors) >= 10:  # Limit error messages
                                    break
            
            # Show results
            if errors:
                QMessageBox.warning(
                    self, "Import Complete with Errors",
                    f"Imported {imported} records.\n\n"
                    f"Errors: {len(errors)}\n" + "\n".join(errors[:5])
                )
            else:
                QMessageBox.information(
                    self, "Import Complete",
                    f"Successfully imported {imported} records."
                )
            
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "Import Error", f"Failed to import:\n{str(e)}")
