"""Main application window"""

from pathlib import Path
from typing import Any

from PySide6.QtCore import QEvent, QSize, Qt
from PySide6.QtGui import QAction, QIcon, QKeySequence, QMouseEvent, QPixmap
from PySide6.QtWidgets import (
    QAbstractItemView,
    QFrame,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QSizePolicy,
    QSplitter,
    QToolBar,
    QVBoxLayout,
    QWidget,
)

from src.core.collection_store import CollectionStore
from src.core.config import Config
from src.core.resource_path import asset_path, get_quartz_icon_path
from src.core.update_checker import UpdateCheckWorker
from src.core.version import VERSION
from src.core.workspace import Workspace
from src.ui.advanced_search_dialog import AdvancedSearchDialog
from src.ui.form_view import FormView
from src.ui.sanitize_dialog import SanitizeDialog
from src.ui.styles import AppStyles
from src.ui.table_view import TableView
from src.ui.update_dialog import UpdateDialog
from src.ui.update_progress_dialog import UpdateProgressDialog


class CollectionsListWidget(QListWidget):
    """Custom QListWidget that prevents selection on right-click and supports drag and drop"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._right_click_selected_row = -1
        # Enable drag and drop for reordering
        # Use InternalMove but ensure clicks are processed immediately
        self.setDragDropMode(QAbstractItemView.InternalMove)
        self.setDefaultDropAction(Qt.MoveAction)
        self.setDragDropOverwriteMode(False)
        # Set movement threshold - higher value means clicks register faster
        # Qt's default is usually 4-10 pixels, we'll use a reasonable default

    def mousePressEvent(self, event: QMouseEvent):
        """Override to prevent selection on right-click"""
        if event.button() == Qt.RightButton:
            # Store current selection before processing
            current_item = self.currentItem()
            if current_item:
                self._right_click_selected_row = self.currentRow()
            else:
                self._right_click_selected_row = -1

            # Call parent to show context menu, but block selection
            self.blockSignals(True)
            super().mousePressEvent(event)

            # Immediately restore selection
            if self._right_click_selected_row >= 0 and self._right_click_selected_row < self.count():
                self.setCurrentRow(self._right_click_selected_row)

            self.blockSignals(False)
        else:
            # For left clicks, process immediately - Qt's InternalMove will handle drag detection
            # without delaying clicks
            super().mousePressEvent(event)

    def startDrag(self, supportedActions):
        """Override to only start drag after actual mouse movement"""
        # Let Qt handle the drag start - it already checks for movement
        # This override ensures clicks are processed immediately
        super().startDrag(supportedActions)

    def dropEvent(self, event):
        """Handle drop event to save new collection order"""
        super().dropEvent(event)
        # Notify parent to save the new order (deferred to avoid blocking)
        if hasattr(self.parent(), '_on_collections_reordered'):
            self.parent()._on_collections_reordered()


class MainWindow(QMainWindow):
    """Main application window"""

    def __init__(self, config: Config):
        super().__init__()
        self.config = config
        self.workspace = Workspace(config.workspace_path)
        self.current_collection: str | None = None
        self.current_store: CollectionStore | None = None

        # Store filters and sorting per collection
        self.collection_filters: dict[str, str] = {}  # collection_name -> search_query
        self.collection_sorting: dict[str, tuple] = {}  # collection_name -> (column, order)

        # Active filters (list of filter dicts: {field/text, operator, value})
        self.active_filters: list[dict] = []

        # Form lock state (starts unlocked)
        self.form_locked = False

        # Undo/Redo history
        self.undo_history: list = []  # List of commands that can be undone
        self.redo_history: list = []  # List of commands that can be redone
        self.max_history = 50  # Maximum number of undo/redo steps

        # Update check threads (for proper lifecycle management)
        self.update_check_threads: list = []  # Keep references to prevent garbage collection

        # Icon cache to avoid reloading icons on every refresh
        self._icon_cache: dict[str, QIcon] = {}

        # Set window icon
        icon_path = get_quartz_icon_path()
        if icon_path.exists():
            self.setWindowIcon(QIcon(str(icon_path)))

        # Set window title with version
        self.setWindowTitle(f"Quartz v{VERSION}")

        self._apply_theme()
        self._init_ui()
        self._load_collections()

        # Apply initial view settings
        self._apply_view_settings()

        # Check for updates on startup if enabled
        if self.config.get("auto_check_for_updates", False):
            from PySide6.QtCore import QTimer
            # Delay check slightly to let UI finish loading
            QTimer.singleShot(2000, self._check_for_updates_async)

    def _apply_theme(self):
        """Apply theme stylesheet"""
        # Support new theme system with color_scheme and mode
        color_scheme = self.config.get("color_scheme", "default")

        # Check if mode is explicitly set in config, otherwise use default or migrate from old theme
        if "mode" in self.config.data:
            mode = self.config.get("mode", "light")
        else:
            # Backward compatibility: migrate old theme setting to mode if mode not explicitly set
            old_theme = self.config.get("theme", None)
            if old_theme and old_theme in ["light", "dark", "system"]:
                mode = old_theme if old_theme != "system" else "light"
                # Migrate to new setting
                self.config.set("mode", mode)
            else:
                mode = "light"

        stylesheet = AppStyles.get_theme(color_scheme=color_scheme, mode=mode)
        self.setStyleSheet(stylesheet)

    def _init_ui(self):
        """Initialize UI components"""
        self.setWindowTitle("Quartz")
        self.setMinimumSize(1000, 700)
        self.resize(1200, 500)

        # Central widget with splitter
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        layout = QHBoxLayout(central_widget)
        layout.setContentsMargins(0, 0, 0, 0)

        splitter = QSplitter(Qt.Horizontal)
        layout.addWidget(splitter)

        # Left sidebar - Collections
        sidebar_widget = QWidget()
        self.sidebar_widget = sidebar_widget  # Store reference for visibility toggle
        sidebar_layout = QVBoxLayout(sidebar_widget)
        sidebar_layout.setContentsMargins(0, 0, 0, 0)

        # Sidebar header
        sidebar_header = QHBoxLayout()

        # Add collection button (left side)
        add_collection_btn = QPushButton()
        add_collection_btn.setIcon(QIcon(str(asset_path("create_collection.png"))))
        add_collection_btn.setMaximumWidth(28)
        add_collection_btn.setMaximumHeight(28)
        add_collection_btn.setProperty("class", "icon-button")
        add_collection_btn.setToolTip("New Collection")
        add_collection_btn.clicked.connect(self._new_collection)
        sidebar_header.addWidget(add_collection_btn)

        # Delete collection button (left side)
        self.delete_collection_btn = QPushButton()
        self.delete_collection_btn.setIcon(QIcon(str(asset_path("delete_collection.png"))))
        self.delete_collection_btn.setMaximumWidth(28)
        self.delete_collection_btn.setMaximumHeight(28)
        self.delete_collection_btn.setProperty("class", "icon-button")
        self.delete_collection_btn.setToolTip("Delete Selected Collection")
        self.delete_collection_btn.setEnabled(
            False
        )  # Disabled until collection selected
        self.delete_collection_btn.clicked.connect(self._delete_selected_collection)
        sidebar_header.addWidget(self.delete_collection_btn)

        # Add stretch to push Collections label to center
        sidebar_header.addStretch()

        # Collections label (centered and bold)
        collections_label = QLabel("Collections")
        collections_label.setAlignment(Qt.AlignCenter)
        font = collections_label.font()
        font.setBold(True)
        collections_label.setFont(font)
        sidebar_header.addWidget(collections_label)

        # Add stretch on the right to keep it centered
        sidebar_header.addStretch()

        sidebar_layout.addLayout(sidebar_header)

        self.collections_list = CollectionsListWidget(self)
        self.collections_list.setContextMenuPolicy(Qt.CustomContextMenu)
        self.collections_list.setIconSize(QSize(32, 32))  # Icon size for collections
        self.collections_list.customContextMenuRequested.connect(
            self._show_collection_context_menu
        )
        # Use currentItemChanged instead of itemClicked to avoid drag detection delay
        self.collections_list.currentItemChanged.connect(self._on_collection_selected)
        # Handle clicks on empty space to deselect
        self.collections_list.itemSelectionChanged.connect(self._on_collection_selection_changed)
        # Track if we're in a right-click to prevent selection changes
        self._right_click_in_progress = False
        # Install event filter to detect clicks on empty space in collections list
        self.collections_list.installEventFilter(self)
        sidebar_layout.addWidget(self.collections_list)

        splitter.addWidget(sidebar_widget)
        splitter.setSizes([250, 750])  # Set initial sizes

        # Right side - Main content
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_layout.setContentsMargins(0, 0, 0, 0)

        # Top toolbar
        toolbar_container = QWidget()
        toolbar_layout = QVBoxLayout(toolbar_container)
        toolbar_layout.setContentsMargins(0, 0, 0, 0)
        self._create_toolbar(toolbar_layout)
        right_layout.addWidget(toolbar_container)

        # View toggle and search
        self.top_bar_widget = QWidget()
        top_bar = QHBoxLayout(self.top_bar_widget)
        top_bar.setContentsMargins(8, 8, 8, 8)
        top_bar.setSpacing(12)

        # View toggle (Form/Table) - proper toggle group
        toggle_container = QWidget()
        toggle_layout = QHBoxLayout(toggle_container)
        toggle_layout.setContentsMargins(0, 0, 0, 0)
        toggle_layout.setSpacing(0)

        from PySide6.QtWidgets import QButtonGroup

        self.view_toggle_group = QButtonGroup()

        # Form first, then Table
        self.form_toggle = QPushButton("FORM")
        self.form_toggle.setCheckable(True)
        self.form_toggle.setChecked(True)  # Start with Form view
        self.form_toggle.setProperty("class", "toggle")
        self.form_toggle.clicked.connect(lambda: self._switch_to_view(1))
        self.view_toggle_group.addButton(self.form_toggle, 1)
        toggle_layout.addWidget(self.form_toggle)

        self.table_toggle = QPushButton("TABLE")
        self.table_toggle.setCheckable(True)
        self.table_toggle.setProperty("class", "toggle")
        self.table_toggle.clicked.connect(lambda: self._switch_to_view(0))
        self.view_toggle_group.addButton(self.table_toggle, 0)
        toggle_layout.addWidget(self.table_toggle)

        top_bar.addWidget(toggle_container)

        top_bar.addStretch()

        # Record navigation
        nav_container = QWidget()
        nav_layout = QHBoxLayout(nav_container)
        nav_layout.setContentsMargins(0, 0, 0, 0)
        nav_layout.setSpacing(4)

        self.nav_label = QLabel("No collection")
        nav_layout.addWidget(self.nav_label)

        self.prev_btn = QPushButton()
        self.prev_btn.setIcon(QIcon(str(asset_path("up.png"))))
        self.prev_btn.setProperty("class", "nav")
        self.prev_btn.setFixedSize(20, 20)
        self.prev_btn.setIconSize(QSize(16, 16))  # Slightly smaller than button to ensure it fits
        self.prev_btn.clicked.connect(self._prev_record)
        nav_layout.addWidget(self.prev_btn)

        self.next_btn = QPushButton()
        self.next_btn.setIcon(QIcon(str(asset_path("down.png"))))
        self.next_btn.setProperty("class", "nav")
        self.next_btn.setFixedSize(20, 20)
        self.next_btn.setIconSize(QSize(16, 16))  # Slightly smaller than button to ensure it fits
        self.next_btn.clicked.connect(self._next_record)
        nav_layout.addWidget(self.next_btn)

        top_bar.addWidget(nav_container)

        # Search box
        self.search_box = QLineEdit()
        self.search_box.setProperty("class", "search")
        self.search_box.setPlaceholderText("Search...")
        self.search_box.setToolTip("Type to search and filter records in real-time")
        self.search_box.textChanged.connect(self._on_search)
        self.search_box.setMinimumWidth(250)
        top_bar.addWidget(self.search_box)

        # Filter button
        filter_icon_path = asset_path("filter.png")
        self.filter_btn = QPushButton()
        if filter_icon_path.exists():
            self.filter_btn.setIcon(QIcon(str(filter_icon_path)))
        self.filter_btn.setProperty("class", "nav")
        self.filter_btn.setToolTip("Add Filter")
        self.filter_btn.setFixedSize(20, 20)
        self.filter_btn.setIconSize(QSize(16, 16))
        self.filter_btn.clicked.connect(self._open_filter_dialog)
        top_bar.addWidget(self.filter_btn)

        # Advanced search button
        adv_icon_path = asset_path("adv.png")
        self.adv_search_btn = QPushButton()
        if adv_icon_path.exists():
            self.adv_search_btn.setIcon(QIcon(str(adv_icon_path)))
        self.adv_search_btn.setProperty("class", "nav")
        self.adv_search_btn.setToolTip("Advanced Search (SQL operators, search all collections)")
        self.adv_search_btn.setFixedSize(20, 20)
        self.adv_search_btn.setIconSize(QSize(16, 16))  # Slightly smaller than button to ensure it fits
        self.adv_search_btn.clicked.connect(self._open_advanced_search)
        top_bar.addWidget(self.adv_search_btn)

        # Sanitize button
        broom_icon_path = asset_path("broom.png")
        self.sanitize_btn = QPushButton()
        if broom_icon_path.exists():
            self.sanitize_btn.setIcon(QIcon(str(broom_icon_path)))
        self.sanitize_btn.setProperty("class", "nav")
        self.sanitize_btn.setToolTip("Sanitize (find and merge duplicate records)")
        self.sanitize_btn.setFixedSize(20, 20)
        self.sanitize_btn.setIconSize(QSize(16, 16))
        self.sanitize_btn.clicked.connect(self._open_sanitize_dialog)
        top_bar.addWidget(self.sanitize_btn)

        right_layout.addWidget(self.top_bar_widget)

        # Filter chips container (below search bar)

        self.filter_chips_container = QWidget()
        self.filter_chips_layout = QHBoxLayout(self.filter_chips_container)
        self.filter_chips_layout.setContentsMargins(8, 4, 8, 4)
        self.filter_chips_layout.setSpacing(8)
        self.filter_chips_layout.addStretch()
        self.filter_chips_container.setVisible(False)  # Hidden until filters are added
        # Allow wrapping of filter chips
        self.filter_chips_container.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Minimum)
        right_layout.addWidget(self.filter_chips_container)

        # Install event filter on top_bar_widget to detect clicks on empty space
        self.top_bar_widget.installEventFilter(self)

        # Main content area - Stacked widget for Form/Table
        from PySide6.QtWidgets import QStackedWidget

        self.content_stack = QStackedWidget()

        # Empty state placeholder (shown when no collection is selected)
        self.empty_state_widget = QWidget()
        empty_layout = QVBoxLayout(self.empty_state_widget)
        empty_layout.setAlignment(Qt.AlignCenter)
        empty_label = QLabel("Please select a collection or create new")
        empty_label.setAlignment(Qt.AlignCenter)
        empty_label.setStyleSheet("""
            QLabel {
                font-size: 18px;
                color: #666;
                padding: 40px;
            }
        """)
        empty_layout.addWidget(empty_label)
        self.content_stack.addWidget(self.empty_state_widget)

        # Table view
        self.table_view = TableView()
        self.content_stack.addWidget(self.table_view)

        # Form view
        self.form_view = FormView()
        self.form_view.record_saved.connect(self._on_record_saved)
        self.content_stack.addWidget(self.form_view)

        # Start with empty state
        self.content_stack.setCurrentIndex(0)

        right_layout.addWidget(self.content_stack)

        splitter.addWidget(right_widget)
        splitter.setStretchFactor(1, 1)

        # Menu bar
        self._create_menu_bar()

        # Status bar
        status_bar = self.statusBar()
        status_bar.showMessage("Ready")

        # Add version label to status bar (right side)
        version_label = QLabel(f"v{VERSION}")
        version_label.setStyleSheet("""
            QLabel {
                color: rgba(255, 255, 255, 0.7);
                font-size: 11px;
                padding: 0px 8px;
            }
        """)
        status_bar.addPermanentWidget(version_label)

    def _create_toolbar(self, parent_layout):
        """Create main toolbar"""
        toolbar = QToolBar()
        self.main_toolbar = toolbar  # Store reference for compact view
        toolbar.setMovable(False)
        toolbar.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)

        # Collection actions with icons
        new_record_action = QAction(self)
        new_record_action.setIcon(QIcon(str(asset_path("add_row.png"))))
        new_record_action.setToolTip("New Record (Ctrl+N)")
        new_record_action.setShortcut(QKeySequence("Ctrl+N"))
        new_record_action.triggered.connect(self._new_record)
        self.new_record_action = new_record_action  # Store reference for compact view
        toolbar.addAction(new_record_action)
        # Store widget reference for compact view
        self.new_record_widget = toolbar.widgetForAction(new_record_action)

        delete_record_action = QAction(self)
        delete_record_action.setIcon(QIcon(str(asset_path("delete_row.png"))))
        delete_record_action.setToolTip("Delete Record (Delete)")
        delete_record_action.setShortcut(QKeySequence("Delete"))
        delete_record_action.triggered.connect(self._delete_record)
        self.delete_record_action = delete_record_action  # Store reference for compact view
        toolbar.addAction(delete_record_action)
        # Store widget reference for compact view
        self.delete_record_widget = toolbar.widgetForAction(delete_record_action)

        toolbar.addSeparator()

        # Field actions
        add_field_action = QAction(self)
        add_field_action.setIcon(QIcon(str(asset_path("add_field.png"))))
        add_field_action.setToolTip("Add Field (Ctrl+G)")
        add_field_action.setShortcut(QKeySequence("Ctrl+G"))
        add_field_action.triggered.connect(self._add_field)
        add_field_action.setEnabled(False)  # Disabled until collection selected
        self.add_field_action = add_field_action  # Store reference
        toolbar.addAction(add_field_action)
        # Store widget reference for compact view
        self.add_field_widget = toolbar.widgetForAction(add_field_action)

        delete_field_action = QAction(self)
        delete_field_action.setIcon(QIcon(str(asset_path("delete_field.png"))))
        delete_field_action.setToolTip("Delete Field")
        delete_field_action.triggered.connect(self._delete_field)
        delete_field_action.setEnabled(False)  # Disabled until collection selected
        self.delete_field_action = delete_field_action  # Store reference
        toolbar.addAction(delete_field_action)
        # Store widget reference for compact view
        self.delete_field_widget = toolbar.widgetForAction(delete_field_action)

        toolbar.addSeparator()

        # Undo/Redo actions
        undo_action = QAction(self)
        undo_action.setIcon(QIcon(str(asset_path("undo.png"))))
        undo_action.setToolTip("Undo (Ctrl+Z)")
        undo_action.setShortcut(QKeySequence("Ctrl+Z"))
        undo_action.triggered.connect(self._undo)
        undo_action.setEnabled(False)  # Will be enabled when there's history
        self.undo_action = undo_action
        toolbar.addAction(undo_action)

        redo_action = QAction(self)
        redo_action.setIcon(QIcon(str(asset_path("redo.png"))))
        redo_action.setToolTip("Redo (Ctrl+Y)")
        redo_action.setShortcut(QKeySequence("Ctrl+Y"))
        redo_action.triggered.connect(self._redo)
        redo_action.setEnabled(False)  # Will be enabled when there's redo history
        self.redo_action = redo_action
        toolbar.addAction(redo_action)

        toolbar.addSeparator()

        bulk_add_action = QAction(self)
        bulk_add_action.setIcon(QIcon(str(asset_path("bulk.png"))))
        bulk_add_action.setToolTip("Bulk Add...")
        bulk_add_action.triggered.connect(self._bulk_add_records)
        self.bulk_add_action = bulk_add_action  # Store reference for compact view
        toolbar.addAction(bulk_add_action)

        duplicate_record_action = QAction(self)
        duplicate_record_action.setIcon(QIcon(str(asset_path("duplicate.png"))))
        duplicate_record_action.setToolTip("Duplicate Record")
        duplicate_record_action.triggered.connect(self._duplicate_record)
        self.duplicate_record_action = duplicate_record_action  # Store reference for compact view
        toolbar.addAction(duplicate_record_action)

        toolbar.addSeparator()

        # Data actions
        import_action = QAction("Import", self)
        import_action.setIcon(QIcon(str(asset_path("import.png"))))
        import_action.triggered.connect(self._import_data)
        import_action.setEnabled(False)  # Disabled until collection selected
        self.import_action = import_action  # Store reference to enable/disable
        toolbar.addAction(import_action)

        export_action = QAction("Export", self)
        export_action.setIcon(QIcon(str(asset_path("export.png"))))
        export_action.triggered.connect(self._export_data)
        toolbar.addAction(export_action)

        toolbar.addSeparator()

        # Upload action (create new collection from CSV)
        upload_action = QAction("Upload", self)
        upload_action.setIcon(QIcon(str(asset_path("upload.png"))))
        upload_action.triggered.connect(self._upload_data)
        toolbar.addAction(upload_action)

        toolbar.addSeparator()

        # Join action (relational join)
        join_action = QAction(self)
        join_action.setIcon(QIcon(str(asset_path("join.png"))))
        join_action.setToolTip("Relational Join...")
        join_action.triggered.connect(self._show_join_dialog)
        join_action.setEnabled(False)  # Disabled until collection selected
        self.join_action = join_action  # Store reference for compact view
        toolbar.addAction(join_action)
        # Store widget reference for compact view
        self.join_widget = toolbar.widgetForAction(join_action)

        toolbar.addSeparator()

        # Lock/Unlock form toggle
        self.lock_form_action = QAction(self)
        self.lock_form_action.setIcon(QIcon(str(asset_path("unlock.png"))))
        self.lock_form_action.setToolTip("Lock/Unlock Form (Toggle Readonly)")
        self.lock_form_action.setCheckable(True)
        self.lock_form_action.setChecked(False)  # Start unlocked
        self.lock_form_action.triggered.connect(self._toggle_form_lock)
        toolbar.addAction(self.lock_form_action)
        # Store widget reference for compact view
        self.lock_form_widget = toolbar.widgetForAction(self.lock_form_action)

        # Apply compact view settings (after all widgets are created)
        self._update_compact_view()

        toolbar.addSeparator()

        # Settings
        settings_action = QAction(self)
        settings_action.setIcon(QIcon(str(asset_path("settings.png"))))
        settings_action.setToolTip("Settings")
        settings_action.triggered.connect(self._show_settings)
        toolbar.addAction(settings_action)

        # Overflow indicator (shows when toolbar items are hidden due to window size)
        from PySide6.QtWidgets import QMenu, QToolButton
        self.overflow_button = QToolButton()
        self.overflow_button.setText("⋯")  # Horizontal ellipsis
        self.overflow_button.setToolTip("More options (hidden toolbar items)")
        self.overflow_button.setPopupMode(QToolButton.InstantPopup)
        self.overflow_button.setVisible(False)  # Hidden by default
        self.overflow_button.setStyleSheet("""
            QToolButton {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #9c27b0, stop:1 #7b1fa2);
                color: white;
                border: 1px solid #7b1fa2;
                border-radius: 4px;
                padding: 4px 8px;
                font-size: 14px;
                font-weight: bold;
                min-width: 24px;
            }
            QToolButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1, stop:0 #ab47bc, stop:1 #8e24aa);
            }
        """)
        self.overflow_menu = QMenu(self.overflow_button)
        self.overflow_button.setMenu(self.overflow_menu)
        toolbar.addWidget(self.overflow_button)

        parent_layout.addWidget(toolbar)

        # Check overflow after window is shown and on layout changes
        from PySide6.QtCore import QTimer
        QTimer.singleShot(300, self._check_toolbar_overflow)

    def _create_menu_bar(self):
        """Create menu bar"""
        menubar = self.menuBar()

        # File menu
        file_menu = menubar.addMenu("File")

        new_collection_action = QAction("New Collection...", self)
        new_collection_action.setShortcut(QKeySequence("Ctrl+Shift+N"))
        new_collection_action.triggered.connect(self._new_collection)
        file_menu.addAction(new_collection_action)

        file_menu.addSeparator()

        export_all_action = QAction("Export All Collections...", self)
        export_all_action.triggered.connect(self._export_all_collections)
        file_menu.addAction(export_all_action)

        file_menu.addSeparator()

        delete_all_records_action = QAction("Delete All Records...", self)
        delete_all_records_action.triggered.connect(self._delete_all_records)
        file_menu.addAction(delete_all_records_action)

        delete_all_collections_action = QAction("Delete All Collections...", self)
        delete_all_collections_action.triggered.connect(self._delete_all_collections)
        file_menu.addAction(delete_all_collections_action)

        file_menu.addSeparator()

        exit_action = QAction("Exit", self)
        exit_action.setShortcut(QKeySequence("Ctrl+Q"))
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        # View menu
        view_menu = menubar.addMenu("View")

        compact_view_action = QAction("Compact View", self)
        compact_view_action.setCheckable(True)
        compact_view_action.setChecked(self.config.get("compact_view", False))
        compact_view_action.triggered.connect(self._toggle_compact_view)
        self.compact_view_action = compact_view_action
        view_menu.addAction(compact_view_action)

        visible_collection_panel_action = QAction("Visible Collection Panel", self)
        visible_collection_panel_action.setCheckable(True)
        visible_collection_panel_action.setChecked(self.config.get("visible_collection_panel", True))
        visible_collection_panel_action.triggered.connect(self._toggle_collection_panel)
        self.visible_collection_panel_action = visible_collection_panel_action
        view_menu.addAction(visible_collection_panel_action)

        view_menu.addSeparator()

        show_key_action = QAction("Show Key", self)
        show_key_action.setCheckable(True)
        show_key_action.setChecked(self.config.get("show_key_column", True))
        show_key_action.triggered.connect(self._toggle_show_key)
        self.show_key_action = show_key_action
        view_menu.addAction(show_key_action)

        view_menu.addSeparator()

        expanded_view_action = QAction("Expanded View", self)
        expanded_view_action.setCheckable(True)
        expanded_view_action.setChecked(self.config.get("expanded_view", False))
        expanded_view_action.triggered.connect(self._toggle_expanded_view)
        self.expanded_view_action = expanded_view_action
        view_menu.addAction(expanded_view_action)

        # Tools menu
        tools_menu = menubar.addMenu("Tools")

        refresh_action = QAction("Refresh", self)
        refresh_action.setShortcut(QKeySequence("F5"))
        refresh_action.triggered.connect(self._refresh_all)
        tools_menu.addAction(refresh_action)

        tools_menu.addSeparator()

        audit_trail_action = QAction("Audit Trail...", self)
        audit_trail_action.setShortcut(QKeySequence("Ctrl+Shift+Z"))
        audit_trail_action.triggered.connect(self._show_audit_trail)
        tools_menu.addAction(audit_trail_action)

        tools_menu.addSeparator()

        shortcuts_action = QAction("Shortcuts...", self)
        shortcuts_action.triggered.connect(self._show_shortcuts)
        tools_menu.addAction(shortcuts_action)

        preferences_action = QAction("Preferences...", self)
        preferences_action.triggered.connect(self._show_settings)
        tools_menu.addAction(preferences_action)

        tools_menu.addSeparator()

        check_update_action = QAction("Check for Updates...", self)
        check_update_action.triggered.connect(self._manual_check_for_updates)
        tools_menu.addAction(check_update_action)

        # Placeholder action for Ctrl+F (currently does nothing)
        placeholder_action = QAction("Search (Placeholder)", self)
        placeholder_action.setShortcut(QKeySequence("Ctrl+F"))
        placeholder_action.setEnabled(False)  # Disabled as placeholder
        tools_menu.addAction(placeholder_action)

    def _load_collections(self):
        """Load collections into sidebar"""
        self.collections_list.clear()
        for name in self.workspace.list_collections():
            item = QListWidgetItem(name)

            # Load collection icon if available (with caching)
            icon_path = self.workspace.get_collection_icon_path(name)
            cache_key = f"collection_{name}"

            if cache_key not in self._icon_cache:
                if icon_path and icon_path.exists():
                    pixmap = QPixmap(str(icon_path))
                    if not pixmap.isNull():
                        # Scale to icon size
                        scaled_pixmap = pixmap.scaled(
                            32, 32, Qt.KeepAspectRatio, Qt.SmoothTransformation
                        )
                        self._icon_cache[cache_key] = QIcon(scaled_pixmap)
                else:
                    # Default icon (quartz crystal) - cache this too
                    if "default" not in self._icon_cache:
                        default_icon_path = get_quartz_icon_path()
                        if default_icon_path.exists():
                            pixmap = QPixmap(str(default_icon_path))
                            if not pixmap.isNull():
                                scaled_pixmap = pixmap.scaled(
                                    32, 32, Qt.KeepAspectRatio, Qt.SmoothTransformation
                                )
                                self._icon_cache["default"] = QIcon(scaled_pixmap)
                    if "default" in self._icon_cache:
                        self._icon_cache[cache_key] = self._icon_cache["default"]

            if cache_key in self._icon_cache:
                item.setIcon(self._icon_cache[cache_key])

            self.collections_list.addItem(item)

    def _on_collections_reordered(self):
        """Handle collection reordering via drag and drop"""
        # Get current order of collections from the list widget
        collection_names = []
        for i in range(self.collections_list.count()):
            item = self.collections_list.item(i)
            if item:
                collection_names.append(item.text())

        # Save the new order to workspace (async to avoid blocking UI)
        if collection_names:
            from PySide6.QtCore import QTimer
            # Use a single-shot timer to defer the save operation
            QTimer.singleShot(0, lambda: self.workspace.set_collection_order(collection_names))

    def _show_collection_context_menu(self, position):
        """Show context menu for collections"""
        item = self.collections_list.itemAt(position)
        if not item:
            return

        from PySide6.QtWidgets import QMenu

        # Use the row stored by eventFilter if available, otherwise get current selection
        if hasattr(self, '_right_click_selected_row') and self._right_click_selected_row >= 0:
            selected_row = self._right_click_selected_row
            current_item = self.collections_list.item(selected_row) if selected_row < self.collections_list.count() else None
        else:
            current_item = self.collections_list.currentItem()
            selected_row = self.collections_list.currentRow() if current_item else -1

        clicked_collection_name = item.text()
        was_currently_selected = (current_item and current_item.text() == clicked_collection_name)

        # Store the actual selected collection name for restoration
        selected_collection_name = current_item.text() if current_item else None

        menu = QMenu(self)
        collection_name = clicked_collection_name

        # Properties
        properties_action = menu.addAction("Properties...")
        properties_action.triggered.connect(
            lambda: self._show_collection_properties(collection_name)
        )

        # Set Icon
        set_icon_action = menu.addAction("Set Icon...")
        set_icon_action.triggered.connect(
            lambda: self._set_collection_icon(collection_name)
        )

        menu.addSeparator()

        # Relational Join
        relational_join_action = menu.addAction("Relational Join...")
        relational_join_action.triggered.connect(
            lambda: self._show_relational_join(collection_name)
        )

        # Merge Into
        merge_action = menu.addAction("Merge Into...")
        merge_action.triggered.connect(
            lambda: self._merge_collection_into(collection_name)
        )

        menu.addSeparator()

        # Import Foreign Keys (only if another collection is selected)
        if self.current_collection and self.current_collection != collection_name:
            import_keys_action = menu.addAction("Import Foreign Keys...")
            import_keys_action.triggered.connect(
                lambda: self._import_foreign_keys(self.current_collection, collection_name)
            )
            menu.addSeparator()

        # Delete
        delete_action = menu.addAction("Delete...")
        delete_action.triggered.connect(lambda: self._delete_collection(collection_name))

        # Block signals temporarily to prevent selection change from triggering events
        self.collections_list.blockSignals(True)

        menu.exec(self.collections_list.mapToGlobal(position))

        # Restore selection if Qt automatically selected a different collection on right-click
        # Use row index for more reliable restoration
        if not was_currently_selected and selected_row >= 0:
            # Restore by row index (more reliable than item reference)
            if selected_row < self.collections_list.count():
                self.collections_list.setCurrentRow(selected_row)
        elif not was_currently_selected and selected_collection_name:
            # Fallback: Find the item by name
            items = self.collections_list.findItems(selected_collection_name, Qt.MatchExactly)
            if items:
                self.collections_list.setCurrentItem(items[0])
            elif current_item:
                # Last resort: stored item reference
                self.collections_list.setCurrentItem(current_item)

        # Unblock signals after restoring selection
        self.collections_list.blockSignals(False)

        # Reset the stored row
        self._right_click_selected_row = -1

    def _rename_collection(self, old_name: str):
        """Rename a collection"""
        from PySide6.QtWidgets import QInputDialog

        new_name, ok = QInputDialog.getText(
            self, "Rename Collection", "New name:", text=old_name
        )
        if ok and new_name and new_name != old_name:
            try:
                self.workspace.rename_collection(old_name, new_name)
                self._load_collections()
                # If this was the current collection, reopen it
                if self.current_collection == old_name:
                    items = self.collections_list.findItems(new_name, Qt.MatchExactly)
                    if items:
                        self.collections_list.setCurrentItem(items[0])
                        self._open_collection(new_name)
            except ValueError as e:
                QMessageBox.warning(self, "Error", str(e))

    def _duplicate_collection(self, name: str):
        """Duplicate a collection"""
        from PySide6.QtWidgets import QInputDialog

        new_name, ok = QInputDialog.getText(
            self, "Duplicate Collection", "New collection name:", text=f"{name} Copy"
        )
        if ok and new_name:
            try:
                self.workspace.duplicate_collection(name, new_name)
                self._load_collections()
            except ValueError as e:
                QMessageBox.warning(self, "Error", str(e))

    def _set_collection_icon(self, name: str):
        """Set collection icon/image"""
        from PySide6.QtWidgets import QFileDialog

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Collection Icon",
            "",
            "Image files (*.png *.jpg *.jpeg *.bmp *.gif);;All files (*)",
        )

        if file_path:
            try:
                self.workspace.set_collection_icon(name, Path(file_path))
                self._load_collections()  # Refresh to show new icon
                QMessageBox.information(self, "Success", "Collection icon updated")
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to set icon: {str(e)}")

    def _remove_collection_icon(self, name: str):
        """Remove collection icon"""
        try:
            self.workspace.set_collection_icon(name, None)
            self._load_collections()  # Refresh to remove icon
            QMessageBox.information(self, "Success", "Collection icon removed")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to remove icon: {str(e)}")

    def _show_join_dialog(self):
        """Show relational join dialog for current collection"""
        if not self.current_collection:
            QMessageBox.information(self, "Info", "Please select a collection first")
            return
        self._show_relational_join(self.current_collection)

    def _show_relational_join(self, collection_name: str):
        """Show relational join dialog"""
        from src.ui.relational_join_dialog import RelationalJoinDialog

        dialog = RelationalJoinDialog(self, source_collection=collection_name, workspace=self.workspace)

        if dialog.exec():
            rel_data = dialog.get_relationship_data()

            try:
                # Get source collection store
                source_info = self.workspace.get_collection_info(rel_data["source_collection"])
                source_db = self.workspace.workspace_path / source_info.db_path
                source_store = CollectionStore(source_db)
                source_store.connect()

                # Add relationship to source collection's database
                source_store.add_relationship(
                    relationship_name=rel_data["name"],
                    source_collection=rel_data["source_collection"],
                    source_field_key=rel_data["source_field_key"],
                    target_collection=rel_data["target_collection"],
                    target_field_key=rel_data["target_field_key"],
                    relationship_type=rel_data["type"],
                    cascade_delete=rel_data["cascade_delete"]
                )

                source_store.close()

                QMessageBox.information(
                    self, "Success",
                    f"Relationship '{rel_data['name']}' created successfully!\n\n"
                    f"Source: {rel_data['source_collection']}.{rel_data['source_field_key']}\n"
                    f"Target: {rel_data['target_collection']}.{rel_data['target_field_key']}\n"
                    f"Type: {rel_data['type']}"
                )
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to create relationship: {str(e)}")

    def _merge_collection_into(self, source_name: str):
        """Merge one collection into another"""
        from PySide6.QtWidgets import QInputDialog

        collections = [c for c in self.workspace.list_collections() if c != source_name]
        if not collections:
            QMessageBox.information(self, "Info", "No other collections available to merge into")
            return

        target_name, ok = QInputDialog.getItem(
            self,
            "Merge Collection",
            f"Select target collection to merge '{source_name}' into:",
            collections,
            0,
            False
        )

        if ok and target_name:
            reply = QMessageBox.question(
                self,
                "Confirm Merge",
                f"Merge '{source_name}' into '{target_name}'?\n\n"
                "This will copy all records and fields from the source collection.\n"
                "The source collection will remain unchanged.",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                try:
                    # Get source store
                    source_info = self.workspace.get_collection_info(source_name)
                    source_db = self.workspace.workspace_path / source_info.db_path
                    source_store = CollectionStore(source_db)
                    source_store.connect()

                    # Get target store
                    target_info = self.workspace.get_collection_info(target_name)
                    target_db = self.workspace.workspace_path / target_info.db_path
                    target_store = CollectionStore(target_db)
                    target_store.connect()

                    # Check for conflicts
                    conflicts = self._detect_merge_conflicts(source_store, target_store)

                    # Show conflict resolution dialog if conflicts exist
                    resolutions = {}
                    if (conflicts.get("field_conflicts") or
                        conflicts.get("field_alias_conflicts") or
                        conflicts.get("record_id_conflicts")):
                        from src.ui.merge_conflict_dialog import MergeConflictDialog
                        conflict_dialog = MergeConflictDialog(self, conflicts)
                        if not conflict_dialog.exec():
                            # User cancelled
                            source_store.close()
                            target_store.close()
                            return
                        resolutions = conflict_dialog.get_resolutions()

                    # Resolve field conflicts and copy fields
                    source_fields = source_store.list_fields()
                    target_fields = {f["key"]: f for f in target_store.list_fields()}
                    field_key_mapping = {}  # Map old key to new key

                    for field in source_fields:
                        field_key = field["key"]
                        original_key = field_key

                        # Check for alias conflict first
                        field.get("alias", field.get("label", ""))
                        resolution_alias_key = f"alias_{field_key}"
                        if resolution_alias_key in resolutions:
                            resolution = resolutions[resolution_alias_key]
                            action = resolution["action"]

                            if action == "Keep Separate (Rename Source)":
                                new_key = resolution.get("new_key", f"{field_key}_merged")
                                field_key_mapping[original_key] = new_key
                                field_key = new_key
                            elif action == "Merge Into Target Field":
                                # Map to target key
                                target_key = resolution["target_key"]
                                field_key_mapping[original_key] = target_key
                                field_key = target_key
                                # Skip adding field, use existing target field
                                if field_key in target_fields:
                                    continue

                        # Check if field conflict exists (same key)
                        if field_key in target_fields:
                            # Field conflict - check resolution
                            resolution_key = f"field_{field_key}"
                            if resolution_key in resolutions:
                                resolution = resolutions[resolution_key]
                                action = resolution["action"]

                                if action == "Skip (Keep Target)":
                                    # Skip this field, don't copy
                                    continue
                                elif action == "Rename Source Field":
                                    # Use new key
                                    new_key = resolution.get("new_key", f"{field_key}_merged")
                                    field_key_mapping[field_key] = new_key
                                    field_key = new_key
                                elif action == "Replace Target Field":
                                    # Remove target field first (would need schema migration)
                                    # For now, just skip and warn
                                    QMessageBox.warning(
                                        self, "Warning",
                                        f"Cannot replace field '{field_key}' - field replacement requires schema migration. "
                                        f"Skipping this field."
                                    )
                                    continue
                            else:
                                # No resolution provided, skip
                                continue

                        # Add field if it doesn't exist or was renamed
                        if field_key not in target_fields:
                            # Use alias if available, otherwise label
                            field_label = field.get("alias", field.get("label", field_key))
                            target_store.add_field(
                                field_key=field_key,
                                field_type=field["type"],
                                label=field_label,
                                required=field.get("required", False),
                                default_value=field.get("default_value"),
                                validation_rules=field.get("validation_rules"),
                                options=field.get("options"),
                                indexed=field.get("indexed", False)
                            )

                    # Copy records with conflict handling
                    source_records = source_store.list_records()
                    target_record_ids = {r["id"] for r in target_store.list_records()}
                    records_merged = 0
                    records_skipped = 0

                    for record in source_records:
                        # Check for ID conflict
                        source_id = record.get("id")
                        if source_id and source_id in target_record_ids:
                            # ID conflict - record will get new ID when added
                            # This is expected and handled automatically
                            pass

                        # Map field keys if needed
                        data = {}
                        for old_key, value in record.items():
                            if old_key in ("id", "record_uuid", "created_at", "updated_at"):
                                continue

                            # Use mapped key if field was renamed
                            new_key = field_key_mapping.get(old_key, old_key)

                            # Only include if field exists in target
                            if new_key in [f["key"] for f in target_store.list_fields()]:
                                data[new_key] = value

                        # Add record (will get new ID automatically)
                        try:
                            target_store.add_record(data)
                            records_merged += 1
                        except Exception:
                            records_skipped += 1
                            # Continue with next record

                    source_store.close()
                    target_store.close()

                    # Refresh if target is current collection
                    if self.current_collection == target_name:
                        self._open_collection(target_name)

                    message = f"Merged {records_merged} record(s) from '{source_name}' into '{target_name}'"
                    if records_skipped > 0:
                        message += f"\n{records_skipped} record(s) were skipped due to errors."
                    QMessageBox.information(self, "Success", message)
                except Exception as e:
                    QMessageBox.warning(self, "Error", f"Failed to merge collections: {str(e)}")

    def _detect_merge_conflicts(self, source_store: CollectionStore, target_store: CollectionStore) -> dict:
        """Detect conflicts between source and target collections"""
        conflicts = {
            "field_conflicts": [],
            "field_alias_conflicts": [],
            "record_id_conflicts": []
        }

        # Check for field conflicts (same key)
        source_fields = {f["key"]: f for f in source_store.list_fields()}
        target_fields = {f["key"]: f for f in target_store.list_fields()}
        target_fields_by_alias = {f.get("alias", f.get("label", "")): f for f in target_store.list_fields()}

        for field_key, source_field in source_fields.items():
            if field_key in target_fields:
                # Field conflict - same key but might have different properties
                target_field = target_fields[field_key]

                # Check if properties differ
                source_alias = source_field.get("alias", source_field.get("label", ""))
                target_alias = target_field.get("alias", target_field.get("label", ""))
                source_type = source_field.get("type", "")
                target_type = target_field.get("type", "")

                conflicts["field_conflicts"].append({
                    "key": field_key,
                    "source": source_field,
                    "target": target_field,
                    "alias_differs": source_alias != target_alias,
                    "type_differs": source_type != target_type
                })
            else:
                # Check for alias conflicts (same alias but different key)
                source_alias = source_field.get("alias", source_field.get("label", ""))
                if source_alias and source_alias in target_fields_by_alias:
                    target_field = target_fields_by_alias[source_alias]
                    conflicts["field_alias_conflicts"].append({
                        "source_key": field_key,
                        "target_key": target_field["key"],
                        "alias": source_alias,
                        "source": source_field,
                        "target": target_field
                    })

        # Check for record ID conflicts
        source_records = source_store.list_records()
        target_record_ids = {r["id"] for r in target_store.list_records()}

        conflicting_ids = []
        for record in source_records:
            record_id = record.get("id")
            if record_id and record_id in target_record_ids:
                conflicting_ids.append(record_id)

        if conflicting_ids:
            conflicts["record_id_conflicts"] = conflicting_ids

        return conflicts

    def _import_foreign_keys(self, target_collection: str, source_collection: str):
        """Import foreign keys from another collection"""
        try:
            # Get source store
            source_info = self.workspace.get_collection_info(source_collection)
            source_db = self.workspace.workspace_path / source_info.db_path
            source_store = CollectionStore(source_db)
            source_store.connect()

            # Get target store
            target_info = self.workspace.get_collection_info(target_collection)
            target_db = self.workspace.workspace_path / target_info.db_path
            target_store = CollectionStore(target_db)
            target_store.connect()

            # Get source fields
            source_fields = source_store.list_fields()

            # Show dialog to select which fields to import as foreign keys
            from PySide6.QtWidgets import QCheckBox, QDialog, QLabel, QPushButton, QVBoxLayout

            dialog = QDialog(self)
            dialog.setWindowTitle(f"Import Foreign Keys from '{source_collection}'")
            dialog.setMinimumWidth(400)
            layout = QVBoxLayout(dialog)

            layout.addWidget(QLabel(f"Select fields from '{source_collection}' to import as foreign keys:"))

            checkboxes = {}
            for field in source_fields:
                field_label = field.get('alias', field.get('label', ''))
                cb = QCheckBox(f"{field_label} ({field['key']}) - {field['type']}")
                checkboxes[field['key']] = (cb, field)
                layout.addWidget(cb)

            button_layout = QHBoxLayout()
            ok_btn = QPushButton("Import")
            cancel_btn = QPushButton("Cancel")
            cancel_btn.setProperty("class", "secondary")
            ok_btn.clicked.connect(dialog.accept)
            cancel_btn.clicked.connect(dialog.reject)
            button_layout.addStretch()
            button_layout.addWidget(cancel_btn)
            button_layout.addWidget(ok_btn)
            layout.addLayout(button_layout)

            if dialog.exec():
                imported = 0
                for field_key, (cb, field) in checkboxes.items():
                    if cb.isChecked():
                        # Create foreign key field in target collection
                        # Use a prefix to avoid conflicts
                        target_key = f"{source_collection}_{field_key}"

                        # Check if field already exists
                        existing_fields = {f["key"]: f for f in target_store.list_fields()}
                        if target_key not in existing_fields:
                            # Get field alias or label
                            field_alias = field.get('alias', field.get('label', ''))
                            target_store.add_field(
                                field_key=target_key,
                                field_type="text",  # Foreign keys stored as text (reference to source collection)
                                label=f"{field_alias} (from {source_collection})",
                                required=False
                            )

                            # Create a relationship automatically
                            relationship_name = f"{source_collection}_{field_key}_to_{target_collection}_{target_key}"
                            try:
                                target_store.add_relationship(
                                    relationship_name=relationship_name,
                                    source_collection=source_collection,
                                    source_field_key=field_key,
                                    target_collection=target_collection,
                                    target_field_key=target_key,
                                    relationship_type="one_to_many",
                                    cascade_delete=False
                                )
                            except Exception:
                                # Relationship might already exist, ignore
                                pass

                            imported += 1

                source_store.close()
                target_store.close()

                # Refresh if target is current collection
                if self.current_collection == target_collection:
                    self._open_collection(target_collection)

                QMessageBox.information(
                    self, "Success",
                    f"Imported {imported} foreign key field(s) from '{source_collection}'"
                )
            else:
                source_store.close()
                target_store.close()

        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to import foreign keys: {str(e)}")

    def _show_collection_properties(self, name: str):
        """Show collection properties dialog"""
        info = self.workspace.get_collection_info(name)
        if not info:
            QMessageBox.warning(self, "Error", f"Collection '{name}' not found")
            return

        # Open collection store
        db_path = self.workspace.workspace_path / info.db_path
        store = CollectionStore(db_path)
        store.connect()

        from src.ui.collection_properties_dialog import CollectionPropertiesDialog

        dialog = CollectionPropertiesDialog(self, store, name, self.workspace)
        if dialog.exec():
            # Refresh if this is the current collection
            if self.current_collection == name:
                fields = store.list_fields()
                self.table_view.set_collection(self.current_store, fields)
                self.form_view.set_collection(self.current_store, fields)

        store.close()

    def _export_collection_db(self, name: str):
        """Export collection database"""
        from PySide6.QtWidgets import QFileDialog

        info = self.workspace.get_collection_info(name)
        if not info:
            return

        db_path = self.workspace.workspace_path / info.db_path
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Database",
            f"{name}.sqlite",
            "SQLite Database (*.sqlite);;All Files (*)",
        )
        if file_path:
            import shutil

            shutil.copy2(db_path, file_path)
            QMessageBox.information(
                self, "Success", f"Database exported to {file_path}"
            )

    def _delete_selected_collection(self):
        """Delete the currently selected collection"""
        current_item = self.collections_list.currentItem()
        if not current_item:
            QMessageBox.information(
                self, "Info", "Please select a collection to delete"
            )
            return
        self._delete_collection(current_item.text())

    def _delete_collection(self, name: str):
        """Delete a collection"""
        reply = QMessageBox.question(
            self,
            "Delete Collection",
            f"Delete collection '{name}'?\n\nA backup will be created.",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            try:
                # Close if this is the current collection
                if self.current_collection == name:
                    if self.current_store:
                        self.current_store.close()
                        # Force garbage collection to release file handle
                        import gc
                        gc.collect()
                    self.current_store = None
                    self.current_collection = None
                    # Clear the table and form views
                    self.table_view.set_collection(None, [])
                    self.form_view.set_collection(None, [])

                # Check if backup is enabled in settings
                from src.core.config import Config

                config = Config()
                backup_enabled = config.get("backup_enabled", True)

                # Process events to ensure connection is fully closed
                from PySide6.QtWidgets import QApplication
                QApplication.processEvents()

                self.workspace.delete_collection(name, backup=backup_enabled)

                # Clear icon cache for deleted collection
                cache_key = f"collection_{name}"
                if cache_key in self._icon_cache:
                    del self._icon_cache[cache_key]

                # Immediately refresh the collections list
                self._load_collections()

                # Force update of the list widget to ensure it refreshes
                self.collections_list.update()
                self.collections_list.repaint()

                # Deselect any collection if the deleted one was selected
                if self.current_collection == name:
                    self._deselect_collection()

                self.statusBar().showMessage(f"Deleted collection: {name}")
            except ValueError as e:
                QMessageBox.warning(self, "Error", str(e))

    def _on_collection_selected(self, item: QListWidgetItem, previous: QListWidgetItem = None):
        """Handle collection selection"""
        if item is None:
            return
        collection_name = item.text()
        self._open_collection(collection_name)
        # Enable delete button
        self.delete_collection_btn.setEnabled(True)

    def _on_collection_selection_changed(self):
        """Handle collection selection changes (including deselection)"""
        # If we're in a right-click, restore the original selection and don't process the change
        if hasattr(self, '_right_click_in_progress') and self._right_click_in_progress:
            if hasattr(self, '_right_click_selected_row') and self._right_click_selected_row >= 0 and self._right_click_selected_row < self.collections_list.count():
                self.collections_list.blockSignals(True)
                self.collections_list.setCurrentRow(self._right_click_selected_row)
                self.collections_list.blockSignals(False)
            return

        # If no item is selected, deselect collection
        if not self.collections_list.currentItem():
            self._deselect_collection()

    def _deselect_collection(self):
        """Deselect the current collection"""
        # Close current store
        if self.current_store:
            self.current_store.close()
        self.current_store = None
        self.current_collection = None

        # Clear views
        self.table_view.model.set_collection(None, [])
        self.form_view.set_collection(None, [])

        # Show empty state
        self.content_stack.setCurrentIndex(0)  # Empty state widget

        # Update UI
        self.setWindowTitle("Quartz")
        self.nav_label.setText("No collection")
        self.delete_collection_btn.setEnabled(False)
        # Disable import action (no collection selected)
        if hasattr(self, 'import_action'):
            self.import_action.setEnabled(False)
        # Disable field actions
        if hasattr(self, 'add_field_action'):
            self.add_field_action.setEnabled(False)
        if hasattr(self, 'delete_field_action'):
            self.delete_field_action.setEnabled(False)
        # Disable join action
        if hasattr(self, 'join_action'):
            self.join_action.setEnabled(False)
        self.statusBar().showMessage("No collection selected")

        # Clear search (but don't clear saved filter - it will be restored when collection is reopened)
        self.search_box.clear()

        # Clear current filter from model
        if hasattr(self, 'table_view') and hasattr(self.table_view, 'model'):
            model = self.table_view.model
            model.filtered_records = []
            model.beginResetModel()
            model.endResetModel()

    def _open_collection(self, name: str):
        """Open a collection"""
        info = self.workspace.get_collection_info(name)
        if not info:
            QMessageBox.warning(self, "Error", f"Collection '{name}' not found")
            return

        # Close previous store and cleanup
        if self.current_store:
            self.current_store.close()
            # Clear table view caches to free memory
            if hasattr(self.table_view, 'model'):
                self.table_view.model._record_cache.clear()
                self.table_view.model._loaded_batches.clear()
                self.table_view.model._formatted_cache.clear()

        # Open new collection
        db_path = self.workspace.workspace_path / info.db_path
        self.current_store = CollectionStore(db_path)
        # Load key prefix from collection info if available
        if hasattr(info, 'key_prefix') and info.key_prefix:
            self.current_store.key_prefix = info.key_prefix
        self.current_store.connect()
        self.current_collection = name

        # Clear active filters when switching collections (filters don't persist)
        self.active_filters = []
        self._update_filter_chips()

        # Update views
        fields = self.current_store.list_fields()
        self.table_view.set_collection(self.current_store, fields)
        self.form_view.set_collection(self.current_store, fields)

        # Switch to table or form view (not empty state)
        # Index 1 = table view, Index 2 = form view
        if hasattr(self, 'form_toggle') and self.form_toggle.isChecked():
            self.content_stack.setCurrentIndex(2)  # Form view
        else:
            self.content_stack.setCurrentIndex(1)  # Table view

        # Apply table view settings
        self._apply_table_view_settings()

        # Apply key column visibility setting
        show_key = self.config.get("show_key_column", True)
        self.table_view.setColumnHidden(0, not show_key)

        # Restore saved filter and sorting for this collection
        if not hasattr(self, 'collection_filters'):
            self.collection_filters: dict[str, str] = {}
            self.collection_sorting: dict[str, tuple] = {}

        saved_query = self.collection_filters.get(name, "")
        if saved_query:
            self.search_box.setText(saved_query)
            self._perform_search()
        else:
            self.search_box.clear()
            # Clear any existing filters
            model = self.table_view.model
            model.filtered_records = model.records.copy()
            model.beginResetModel()
            model.endResetModel()

        # Update UI
        self.setWindowTitle(f"Quartz - {name}")
        self._update_navigation()
        self.statusBar().showMessage(f"Opened collection: {name}")
        # Enable delete button
        self.delete_collection_btn.setEnabled(True)

        # Enable import action (collection is selected)
        if hasattr(self, 'import_action'):
            self.import_action.setEnabled(True)

        # Enable field actions
        if hasattr(self, 'add_field_action'):
            self.add_field_action.setEnabled(True)
        if hasattr(self, 'delete_field_action'):
            self.delete_field_action.setEnabled(True)
        # Enable join action
        if hasattr(self, 'join_action'):
            self.join_action.setEnabled(True)

        # Select first record if available
        if self.table_view.model.rowCount() > 0:
            self.table_view.selectRow(0)
            if self.form_toggle.isChecked():
                self.form_view.new_record()

    def _switch_to_view(self, index: int):
        """Switch to Table (0) or Form (1) view
        Note: content_stack indices: 0=empty state, 1=table, 2=form
        """
        # Only switch if a collection is selected
        if not self.current_store:
            return

        # Map: 0 = table view (stack index 1), 1 = form view (stack index 2)
        stack_index = index + 1  # 0 -> 1 (table), 1 -> 2 (form)
        self.content_stack.setCurrentIndex(stack_index)

        if index == 1:  # Form view (stack index 2)
            self.form_view.new_record()

    def _on_search(self, query: str):
        """Handle search query (real-time as user types)"""
        if not self.current_store:
            return

        # Debounce search to avoid too many queries
        from PySide6.QtCore import QTimer

        if not hasattr(self, "_search_timer"):
            self._search_timer = QTimer()
            self._search_timer.setSingleShot(True)
            self._search_timer.timeout.connect(self._perform_search)

        # Store query for timer callback
        self._pending_search_query = query

        # Restart timer (300ms delay)
        self._search_timer.stop()
        self._search_timer.start(300)

    def _perform_search(self):
        """Perform the actual search with text query and active filters"""
        if not self.current_store or not self.current_collection:
            return

        query = getattr(self, "_pending_search_query", "")
        model = self.table_view.model

        # Save filter for this collection
        if not hasattr(self, 'collection_filters'):
            self.collection_filters: dict[str, str] = {}
        self.collection_filters[self.current_collection] = query

        # Get all records (either all or filtered by text search)
        if not query.strip() and not self.active_filters:
            # Show all records - no filter active
            model._is_filtered = False
            if model._virtualized and model._total_count > 500:
                model._search_query = None
                model._refresh_data()
            else:
                model.filtered_records = model.records.copy() if model.records else []
        else:
            # A search or filter is active
            model._is_filtered = True
            # Use simple search for text query (autofilter as user types)
            if query.strip():
                model.filtered_records = self.current_store.simple_search(query)
            else:
                # No text query, start with all records
                model.filtered_records = model.records.copy() if model.records else []

            # Apply active filters (AND logic - all filters must match)
            if self.active_filters:
                model.filtered_records = self._apply_filters(model.filtered_records)

            # Store search query for virtualized mode
            model._search_query = query

        # Clear caches when filtering to ensure fresh data is displayed
        model._formatted_cache.clear()
        model._record_cache.clear()
        model._loaded_batches.clear()

        model.beginResetModel()
        model.endResetModel()

        # Update navigation
        self._update_navigation()

    def _apply_filters(self, records: list[dict]) -> list[dict]:
        """Apply active filters to records (AND logic)"""
        model = self.table_view.model
        model._filter_error = None  # Clear any previous error

        if not self.active_filters or not records:
            return records

        # Validate filters first
        if not self.current_store:
            model._filter_error = "Cannot display information: No collection selected"
            return []

        fields = self.current_store.list_fields()
        field_keys = {f["key"] for f in fields}

        filtered = records
        for filter_item in self.active_filters:
            field_or_text = filter_item.get("field_or_text")
            operator = filter_item.get("operator")
            value = filter_item.get("value")

            if not field_or_text or not operator:
                model._filter_error = "Cannot display information: Invalid filter configuration"
                return []

            # Validate field exists (unless it's a text search)
            if field_or_text != "text" and field_or_text not in field_keys:
                field_label = filter_item.get("field_label", field_or_text)
                model._filter_error = f"Cannot display information: Field '{field_label}' does not exist"
                return []

            # Validate value is provided (unless operator is IS NULL or IS NOT NULL)
            if operator not in ("IS NULL", "IS NOT NULL") and (value is None or (isinstance(value, str) and not value.strip())):
                model._filter_error = "Cannot display information: Filter value is required"
                return []

            # Filter the records
            new_filtered = []
            for record in filtered:
                if field_or_text == "text":
                    # Search across all fields
                    matches = False
                    for field_key, field_value in record.items():
                        if field_key == "id":
                            continue
                        field_str = str(field_value) if field_value is not None else ""
                        if self._match_filter(field_str, operator, value):
                            matches = True
                            break
                    if matches:
                        new_filtered.append(record)
                else:
                    # Filter by specific field
                    field_value = record.get(field_or_text)

                    # Check if this is a checkbox/boolean field
                    field_info = next((f for f in self.current_store.list_fields() if f["key"] == field_or_text), None)
                    is_checkbox = field_info and field_info.get("type") == "checkbox"

                    if is_checkbox:
                        # Handle checkbox/boolean filtering
                        if self._match_checkbox_filter(field_value, operator, value):
                            new_filtered.append(record)
                    else:
                        # Regular field filtering
                        field_str = str(field_value) if field_value is not None else ""
                        if self._match_filter(field_str, operator, value):
                            new_filtered.append(record)

            filtered = new_filtered

        return filtered

    def _match_filter(self, field_value: str, operator: str, filter_value: str) -> bool:
        """Check if field value matches filter criteria"""
        field_lower = field_value.lower()
        filter_lower = str(filter_value).lower()

        if operator == "=" or operator == "IS":
            # Exact match (case-insensitive for strings, exact for numbers)
            try:
                # Try numeric comparison
                field_num = float(field_value)
                filter_num = float(filter_value)
                return abs(field_num - filter_num) < 0.0001  # Handle floating point precision
            except (ValueError, TypeError):
                return field_lower == filter_lower
        elif operator == "!=" or operator == "IS NOT":
            try:
                field_num = float(field_value)
                filter_num = float(filter_value)
                return abs(field_num - filter_num) >= 0.0001
            except (ValueError, TypeError):
                return field_lower != filter_lower
        elif operator == "LIKE":
            return filter_lower in field_lower
        elif operator == "NOT LIKE":
            return filter_lower not in field_lower
        elif operator == ">":
            try:
                return float(field_value) > float(filter_value)
            except (ValueError, TypeError):
                return False
        elif operator == "<":
            try:
                return float(field_value) < float(filter_value)
            except (ValueError, TypeError):
                return False
        elif operator == ">=":
            try:
                return float(field_value) >= float(filter_value)
            except (ValueError, TypeError):
                return False
        elif operator == "<=":
            try:
                return float(field_value) <= float(filter_value)
            except (ValueError, TypeError):
                return False

        return False

    def _match_checkbox_filter(self, field_value: Any, operator: str, filter_value: str) -> bool:
        """Check if checkbox/boolean field value matches filter criteria"""
        # Normalize field value to boolean
        field_bool = self._normalize_to_bool(field_value)

        # Normalize filter value to boolean
        filter_bool = self._normalize_to_bool(filter_value)

        if operator == "=" or operator == "IS":
            return field_bool == filter_bool
        elif operator == "!=" or operator == "IS NOT":
            return field_bool != filter_bool
        elif operator == "LIKE" or operator == "NOT LIKE":
            # For LIKE operators on checkboxes, treat as substring match on string representation
            field_str = str(field_value).lower() if field_value is not None else ""
            filter_str = str(filter_value).lower()
            if operator == "LIKE":
                return filter_str in field_str
            else:
                return filter_str not in field_str
        else:
            # Comparison operators don't make sense for booleans
            return False

    def _normalize_to_bool(self, value: Any) -> bool:
        """Normalize a value to boolean, accepting various formats"""
        if value is None:
            return False

        # If already boolean
        if isinstance(value, bool):
            return value

        # Convert to string and check common true/false representations
        value_str = str(value).lower().strip()

        # True values
        if value_str in ("true", "1", "yes", "on", "checked", "✓", "☑"):
            return True

        # False values
        if value_str in ("false", "0", "no", "off", "unchecked", "", "✗", "☐"):
            return False

        # Try numeric conversion
        try:
            num = float(value_str)
            return num != 0
        except (ValueError, TypeError):
            pass

        # Default: non-empty string is True
        return bool(value_str)

    def _open_filter_dialog(self):
        """Open filter creation dialog"""
        if not self.current_store:
            return

        from src.ui.filter_dialog import FilterDialog

        dialog = FilterDialog(self, self.current_store)
        if dialog.exec():
            filter_item = dialog.get_filter()
            if filter_item:
                self.active_filters.append(filter_item)
                self._update_filter_chips()
                self._perform_search()

    def _update_filter_chips(self):
        """Update the filter chips display"""
        # Clear existing chips
        while self.filter_chips_layout.count() > 1:  # Keep the stretch
            item = self.filter_chips_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        # Add chips for each active filter
        for i, filter_item in enumerate(self.active_filters):
            chip = self._create_filter_chip(filter_item, i)
            self.filter_chips_layout.insertWidget(self.filter_chips_layout.count() - 1, chip)

        # Show/hide container based on whether there are filters
        self.filter_chips_container.setVisible(len(self.active_filters) > 0)

    def _create_filter_chip(self, filter_item: dict, index: int):
        """Create a filter chip widget"""
        from PySide6.QtWidgets import QLabel

        chip = QFrame()
        chip.setProperty("class", "filter-chip")
        chip_layout = QHBoxLayout(chip)
        chip_layout.setContentsMargins(8, 4, 4, 4)
        chip_layout.setSpacing(6)

        # Build filter text
        field_or_text = filter_item.get("field_or_text", "")
        operator = filter_item.get("operator", "")
        value = filter_item.get("value", "")

        if field_or_text == "text":
            filter_text = f"Text {operator} {value}"
        else:
            # Use field_label if available, otherwise look it up from fields
            field_label = filter_item.get("field_label")
            if not field_label and self.current_store:
                # Look up field label from store
                fields = self.current_store.list_fields()
                field = next((f for f in fields if f["key"] == field_or_text), None)
                if field:
                    field_label = field.get("alias", field.get("label", field_or_text))
                else:
                    field_label = field_or_text  # Fallback to key if not found
            filter_text = f"{field_label} {operator} {value}"

        # Label with ellipsis
        label = QLabel(filter_text)
        label.setProperty("class", "filter-chip-label")
        label.setToolTip(filter_text)  # Full text on hover
        label.setWordWrap(False)
        label.setMaximumWidth(200)  # Max width before ellipsis
        label.setTextFormat(Qt.PlainText)
        # Enable ellipsis
        metrics = label.fontMetrics()
        elided_text = metrics.elidedText(filter_text, Qt.ElideRight, 200)
        label.setText(elided_text)
        chip_layout.addWidget(label)

        # Remove icon - use QLabel with pixmap, not a button
        remove_icon_path = asset_path("removefilter.png")
        remove_label = QLabel()
        if remove_icon_path.exists():
            # Load pixmap and scale to match text height
            pixmap = QPixmap(str(remove_icon_path))
            text_height = label.fontMetrics().height()
            scaled_pixmap = pixmap.scaled(text_height, text_height, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            remove_label.setPixmap(scaled_pixmap)
            remove_label.setFixedSize(text_height, text_height)
        else:
            remove_label.setText("×")  # Fallback to text if icon doesn't exist
            remove_label.setFixedSize(16, 16)
        remove_label.setProperty("class", "filter-chip-remove")
        remove_label.setToolTip("Remove filter")
        remove_label.setCursor(Qt.PointingHandCursor)  # Show hand cursor on hover
        # Make label clickable
        remove_label.mousePressEvent = lambda event, idx=index: self._remove_filter(idx)
        chip_layout.addWidget(remove_label)

        return chip

    def _remove_filter(self, index: int):
        """Remove a filter by index"""
        if 0 <= index < len(self.active_filters):
            self.active_filters.pop(index)
            self._update_filter_chips()
            self._perform_search()

    def _open_advanced_search(self):
        """Open advanced search dialog"""
        dialog = AdvancedSearchDialog(self, self.workspace)
        dialog.exec()

    def _open_sanitize_dialog(self):
        """Open the sanitize dialog for the current collection"""
        if not self.current_store:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.information(self, "Sanitize", "Please open a collection first.")
            return
        fields = self.current_store.list_fields()
        dialog = SanitizeDialog(self, store=self.current_store, fields=fields)
        dialog.exec()
        # Refresh view in case records were merged
        if self.current_collection:
            self._perform_search()

    def _open_collection_and_record(self, collection_name: str, record_id: int):
        """Open a specific collection and navigate to a specific record"""
        # Find and select the collection
        items = self.collections_list.findItems(collection_name, Qt.MatchExactly)
        if items:
            self.collections_list.setCurrentItem(items[0])
            self._open_collection(collection_name)

            # Wait a bit for collection to load, then select the record
            from PySide6.QtCore import QTimer
            timer = QTimer()
            timer.setSingleShot(True)
            timer.timeout.connect(lambda: self._select_record_by_id(record_id))
            timer.start(100)  # 100ms delay

    def _select_record_by_id(self, record_id: int):
        """Select a record by ID in the table view"""
        if not self.current_store or not self.table_view:
            return

        model = self.table_view.model
        for row, record in enumerate(model.filtered_records):
            if record.get("id") == record_id:
                self.table_view.selectRow(row)
                # Scroll to the row
                self.table_view.scrollTo(model.index(row, 0))
                break

    def _new_collection(self):
        """Create a new collection"""
        from src.ui.new_collection_dialog import NewCollectionDialog

        dialog = NewCollectionDialog(self)
        if dialog.exec():
            name = dialog.get_collection_name()
            key_prefix = dialog.get_key_prefix()
            fields = dialog.get_fields()
            try:
                # Create collection with optional key prefix
                db_path = self.workspace.create_collection(name, key_prefix=key_prefix)

                # Add fields to the collection (if any)
                if fields:
                    store = CollectionStore(db_path)
                    store.connect()
                    # Get existing fields to check for conflicts
                    existing_fields = store.list_fields()
                    existing_keys = {f["key"] for f in existing_fields}

                    for field in fields:
                        # Generate unique key if needed
                        from src.ui.add_field_dialog import _generate_unique_field_key
                        field_key = _generate_unique_field_key(field["key"], existing_fields)

                        # Only add if it doesn't already exist
                        if field_key not in existing_keys:
                            store.add_field(
                                field_key=field_key,
                                field_type=field["type"],
                                label=field["label"],
                                required=field.get("required", False),
                                default_value=field.get("default_value"),
                                validation_rules=field.get("validation_rules"),
                                options=field.get("options"),
                                indexed=field.get("indexed", False),
                            )
                            # Add to existing fields list to prevent future conflicts
                            existing_fields.append({"key": field_key})
                            existing_keys.add(field_key)
                    store.close()

                self._load_collections()
                # Select the new collection
                items = self.collections_list.findItems(name, Qt.MatchExactly)
                if items:
                    self.collections_list.setCurrentItem(items[0])
                    self._open_collection(name)
            except ValueError as e:
                QMessageBox.warning(self, "Error", str(e))

    def _new_record(self):
        """Create a new record"""
        if not self.current_store:
            QMessageBox.information(self, "Info", "Please select a collection first")
            return

        # If in form view, create new record there
        if self.form_toggle.isChecked():
            self.form_view.new_record()
            # Focus first field
            if self.form_view.field_widgets:
                first_widget = next(iter(self.form_view.field_widgets.values()))
                first_widget.setFocus()
        else:
            # In table view, create a new record and select it
            fields = self.current_store.list_fields()
            data = {}
            for field in fields:
                if field.get("default_value"):
                    data[field["key"]] = field["default_value"]

            record_id = self.current_store.add_record(data)

            # Refresh views
            self.table_view.model._refresh_data()
            self._update_navigation()

            # Select the new record
            model = self.table_view.model
            for i, record in enumerate(model.filtered_records):
                if record["id"] == record_id:
                    self.table_view.selectRow(i)
                    break

    def _bulk_add_records(self):
        """Bulk add blank records"""
        if not self.current_store:
            QMessageBox.information(self, "Info", "Please select a collection first")
            return

        from src.ui.bulk_add_dialog import BulkAddDialog

        dialog = BulkAddDialog(self)
        if dialog.exec():
            count = dialog.get_count()

            # Create records
            fields = self.current_store.list_fields()
            created = 0

            for _ in range(count):
                data = {}
                for field in fields:
                    if field.get("default_value"):
                        data[field["key"]] = field["default_value"]

                try:
                    self.current_store.add_record(data)
                    created += 1
                except Exception as e:
                    QMessageBox.warning(
                        self, "Error", f"Failed to create record: {str(e)}"
                    )
                    break

            # Refresh views
            self.table_view.model._refresh_data()
            self._update_navigation()

            QMessageBox.information(
                self, "Success", f"Created {created} blank record(s)."
            )

    def _delete_record(self):
        """Delete selected record(s)"""
        if not self.current_store:
            QMessageBox.information(self, "Info", "Please select a collection first")
            return

        # Get selected rows from table view
        selection = self.table_view.selectionModel().selectedRows()
        if not selection:
            QMessageBox.information(self, "Info", "Please select record(s) to delete")
            return

        # Confirm deletion
        count = len(selection)
        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Delete {count} record(s)?\n\nThis action cannot be undone.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,  # Default to No for safety
        )

        if reply == QMessageBox.Yes:
            # Get record IDs first (before deletion changes indices)
            model = self.table_view.model
            record_ids = []
            for index in selection:
                row = index.row()
                if row < len(model.filtered_records):
                    record = model.filtered_records[row]
                    record_ids.append(record["id"])

            # Delete records
            for record_id in record_ids:
                try:
                    self.current_store.delete_record(record_id)
                except Exception as e:
                    QMessageBox.warning(
                        self, "Error", f"Failed to delete record: {str(e)}"
                    )

            # Save current filter query before refreshing
            current_query = ""
            if hasattr(self, 'collection_filters') and self.current_collection:
                current_query = self.collection_filters.get(self.current_collection, "")
            # Also check model's search query (for virtualized mode)
            if not current_query and hasattr(model, '_search_query') and model._search_query:
                current_query = model._search_query

            # Refresh views
            model._refresh_data()

            # Reapply the filter if one was active
            if current_query and current_query.strip():
                # Set the pending query and perform search to reapply filter
                # Also update search box text to keep UI in sync (block signals to avoid double trigger)
                if hasattr(self, 'search_box'):
                    # Only update if different to avoid unnecessary signal
                    if self.search_box.text() != current_query:
                        self.search_box.blockSignals(True)
                        self.search_box.setText(current_query)
                        self.search_box.blockSignals(False)
                self._pending_search_query = current_query
                self._perform_search()
            elif hasattr(model, '_search_query'):
                # Clear search query if no filter was active
                model._search_query = None
                # Ensure filtered_records shows all records
                if not model._virtualized or model._total_count <= 500:
                    model.filtered_records = model.records.copy() if model.records else []
                model.beginResetModel()
                model.endResetModel()

            self._update_navigation()

            # Clear form view if record was deleted
            if self.form_view.current_record_id in record_ids:
                self.form_view.current_record_id = None
                self.form_view.new_record()

    def _duplicate_record(self):
        """Duplicate selected record"""
        if not self.current_store:
            QMessageBox.information(self, "Info", "Please select a collection first")
            return

        # Get selected record
        selection = self.table_view.selectionModel().selectedRows()
        if not selection:
            QMessageBox.information(self, "Info", "Please select a record to duplicate")
            return

        model = self.table_view.model
        row = selection[0].row()
        if row >= len(model.filtered_records):
            return

        record = model.filtered_records[row]
        record_id = record["id"]

        # Get full record data
        full_record = self.current_store.get_record(record_id)
        if not full_record:
            QMessageBox.warning(self, "Error", "Failed to load record")
            return

        # Remove ID and UUID to create new record
        data = {
            k: v
            for k, v in full_record.items()
            if k not in ["id", "record_uuid", "created_at", "updated_at"]
        }

        try:
            new_id = self.current_store.add_record(data)
            # Refresh views
            self.table_view.model._refresh_data()
            self._update_navigation()

            # Select the new record
            for i, rec in enumerate(model.records):
                if rec["id"] == new_id:
                    model.index(i, 0)
                    self.table_view.selectRow(i)
                    if self.content_stack.currentIndex() == 2:  # Form view
                        self.form_view.load_record(new_id)
                    break

            QMessageBox.information(self, "Success", "Record duplicated")
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to duplicate record: {str(e)}")

    def _import_data(self):
        """Import data from CSV into existing collection"""
        if not self.current_store:
            QMessageBox.information(self, "Info", "Please select a collection first")
            return

        from src.ui.import_dialog import ImportDialog

        fields = self.current_store.list_fields()
        dialog = ImportDialog(self, self.current_store, fields)
        if dialog.exec():
            # Refresh views after import
            self.table_view.model._refresh_data()
            self._update_navigation()

    def _export_data(self):
        """Export data"""
        if not self.current_store:
            QMessageBox.information(self, "Info", "Please select a collection first")
            return

        # Get selected record IDs
        selected_ids = []
        selection = self.table_view.selectionModel().selectedRows()
        if selection:
            model = self.table_view.model
            for index in selection:
                row = index.row()
                # Use _get_record to handle virtualization properly
                record = model._get_record(row)
                if record:
                    record_id = record.get("id")
                    if record_id is not None:
                        selected_ids.append(record_id)

        # Open export dialog
        from src.core.export_service import ExportService
        from src.ui.export_dialog import ExportDialog

        export_service = ExportService(self.current_store)
        dialog = ExportDialog(
            self, export_service, selected_ids if selected_ids else None
        )
        dialog.exec()

    def _add_field(self):
        """Add a new field to the current collection"""
        if not self.current_store or not self.current_collection:
            QMessageBox.information(self, "Info", "Please select a collection first")
            return

        from src.ui.add_field_dialog import AddFieldDialog

        # Get existing fields to show in dialog
        existing_fields = self.current_store.list_fields()
        dialog = AddFieldDialog(self, existing_fields=existing_fields)
        if dialog.exec():
            field_data = dialog.get_field_data()

            try:
                # Add field to collection
                self.current_store.add_field(
                    field_key=field_data["key"],
                    field_type=field_data["type"],
                    label=field_data["label"],
                    required=field_data.get("required", False),
                    indexed=True,  # Default to indexed for searchability
                    options=field_data.get("options")  # Include options for select/dropdown fields
                )

                # Handle image association if provided
                if field_data.get("image_path"):
                    # Store image association (this would need to be implemented in the store)
                    # For now, we'll just note it
                    pass

                # Refresh views
                fields = self.current_store.list_fields()
                self.table_view.set_collection(self.current_store, fields)
                self.form_view.set_collection(self.current_store, fields)

                QMessageBox.information(self, "Success", f"Field '{field_data['label']}' added successfully")
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to add field: {str(e)}")

    def _delete_field(self):
        """Delete a field from the current collection"""
        if not self.current_store or not self.current_collection:
            QMessageBox.information(self, "Info", "Please select a collection first")
            return

        from PySide6.QtWidgets import QInputDialog

        fields = self.current_store.list_fields()
        if not fields:
            QMessageBox.information(self, "Info", "No fields to delete")
            return

        # Create list of field labels for selection
        field_labels = [f.get("alias", f.get("label", f["key"])) for f in fields]

        field_label, ok = QInputDialog.getItem(
            self,
            "Delete Field",
            "Select field to delete:",
            field_labels,
            0,
            False
        )

        if not ok or not field_label:
            return

        # Find the field
        field = next((f for f in fields if f.get("alias", f.get("label", f["key"])) == field_label), None)
        if not field:
            return

        field_key = field["key"]

        # Confirm deletion
        reply = QMessageBox.warning(
            self,
            "Delete Field",
            f"Are you sure you want to delete the field '{field_label}'?\n\n"
            f"⚠️ WARNING:\n"
            f"• All data in this field will be permanently lost\n"
            f"• This action cannot be undone\n"
            f"• Any dependencies or relationships using this field may break\n"
            f"• The column will be removed from all records\n\n"
            f"This is a destructive operation. Consider exporting your data first.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply == QMessageBox.Yes:
            try:
                # Actually delete the field from the collection
                self.current_store.remove_field(field_key)

                # Refresh the collection view to reflect the change
                if self.current_collection:
                    self._open_collection(self.current_collection)

                self.statusBar().showMessage(f"Field '{field_label}' deleted successfully", 3000)
            except Exception as e:
                QMessageBox.warning(self, "Error", f"Failed to delete field: {str(e)}")

    def _export_all_collections(self):
        """Export all collections as a zip file"""
        import zipfile
        from datetime import datetime

        from PySide6.QtWidgets import QFileDialog, QMessageBox

        collections = self.workspace.list_collections()
        if not collections:
            QMessageBox.information(self, "Info", "No collections to export")
            return

        # Get save location
        default_filename = f"quartz_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export All Collections",
            default_filename,
            "ZIP files (*.zip);;All files (*)"
        )

        if not file_path:
            return

        try:
            with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Add all collection databases
                for collection_name in collections:
                    info = self.workspace.get_collection_info(collection_name)
                    if info:
                        db_path = self.workspace.workspace_path / info.db_path
                        if db_path.exists():
                            # Add database with collection name prefix
                            zipf.write(db_path, f"{collection_name}/{db_path.name}")

                        # Add collection icon if exists
                        icon_path = self.workspace.get_collection_icon_path(collection_name)
                        if icon_path and icon_path.exists():
                            zipf.write(icon_path, f"{collection_name}/icon.png")

                        # Add attachments if they exist
                        collection_dir = self.workspace.workspace_path / collection_name
                        attachments_dir = collection_dir / "attachments"
                        if attachments_dir.exists():
                            for att_file in attachments_dir.rglob("*"):
                                if att_file.is_file():
                                    arcname = f"{collection_name}/attachments/{att_file.relative_to(attachments_dir)}"
                                    zipf.write(att_file, arcname)

            QMessageBox.information(
                self,
                "Success",
                f"Exported {len(collections)} collection(s) to:\n{file_path}"
            )
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to export collections: {str(e)}")

    def _upload_data(self):
        """Upload CSV and create a new collection database"""
        import csv
        from pathlib import Path

        from PySide6.QtWidgets import QFileDialog, QInputDialog, QMessageBox

        # Get CSV or Excel file
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Upload CSV or Excel to Create Collection", "",
            "CSV files (*.csv);;Excel files (*.xlsx *.xls);;All files (*)"
        )
        if not file_path:
            return

        # Get collection name
        collection_name, ok = QInputDialog.getText(
            self, "New Collection", "Enter collection name:"
        )
        if not ok or not collection_name.strip():
            return

        collection_name = collection_name.strip()

        # Check if collection already exists
        if collection_name in self.workspace.list_collections():
            QMessageBox.warning(
                self, "Error", f"Collection '{collection_name}' already exists"
            )
            return

        try:
            # Create new collection
            db_path = self.workspace.create_collection(collection_name)
            store = CollectionStore(db_path)
            store.connect()

            # Read CSV or Excel file
            file_path_obj = Path(file_path)
            file_ext = file_path_obj.suffix.lower()

            if file_ext in ('.xlsx', '.xls'):
                # Read Excel file using openpyxl
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path, read_only=True, data_only=True)
                    ws = wb.active

                    # Get headers from first row
                    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                    csv_headers = [str(cell) if cell is not None else f"Column_{i+1}" for i, cell in enumerate(header_row)]

                    # Get all data rows
                    csv_data = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_data = [str(cell) if cell is not None else "" for cell in row]
                        csv_data.append(row_data)

                    wb.close()
                except ImportError:
                    QMessageBox.critical(self, "Error", "openpyxl is required for Excel files. Install it with: pip install openpyxl")
                    store.close()
                    self.workspace.delete_collection(collection_name, backup=False)
                    return
            else:
                # Read CSV file using built-in csv module
                # Detect encoding first
                from src.ui.import_dialog import detect_file_encoding
                encoding = detect_file_encoding(file_path_obj)
                try:
                    with open(file_path, encoding=encoding) as f:
                        reader = csv.reader(f)
                        csv_headers = next(reader)
                        csv_data = list(reader)
                except UnicodeDecodeError:
                    # If detected encoding fails, try with error handling
                    with open(file_path, encoding=encoding, errors='replace') as f:
                        reader = csv.reader(f)
                        csv_headers = next(reader)
                        csv_data = list(reader)

            # Create fields from CSV headers
            fields_created = []
            for header in csv_headers:
                # Generate field key from header
                field_key = header.lower().replace(" ", "_").replace("-", "_")
                field_key = "".join(c for c in field_key if c.isalnum() or c == "_")
                if not field_key or field_key[0].isdigit():
                    field_key = f"field_{field_key}" if field_key else f"field_{len(fields_created)}"

                # Add field
                store.add_field(
                    field_key=field_key,
                    field_type="text",  # Default to text
                    label=header,
                    required=False
                )
                fields_created.append({"key": field_key, "label": header})

            # Import data
            imported = 0
            for row_data in csv_data:
                record_data = {}
                for i, value in enumerate(row_data):
                    if i < len(fields_created) and value:
                        field_key = fields_created[i]["key"]
                        record_data[field_key] = str(value)

                if record_data:
                    store.add_record(record_data)
                    imported += 1

            store.close()

            # Refresh collections list
            self._load_collections()

            # Open the new collection
            self._open_collection(collection_name)

            QMessageBox.information(
                self, "Success",
                f"Collection '{collection_name}' created with {imported} records from CSV."
            )
        except Exception as e:
            QMessageBox.critical(
                self, "Error",
                f"Failed to upload CSV and create collection:\n{str(e)}"
            )

    def _toggle_form_lock(self, checked: bool):
        """Toggle form lock state (readonly/editable)"""
        self.form_locked = checked

        # Update icon (only if not in compact view)
        if not self.config.get("compact_view", False):
            if checked:
                self.lock_form_action.setIcon(QIcon(str(asset_path("lock.png"))))
                self.lock_form_action.setToolTip("Unlock Form (Make Editable)")
            else:
                self.lock_form_action.setIcon(QIcon(str(asset_path("unlock.png"))))
                self.lock_form_action.setToolTip("Lock Form (Make Readonly)")

        # Update form view readonly state
        self.form_view.set_readonly(checked)

        # Update table view readonly state
        self.table_view.set_readonly(checked)

    def _toggle_compact_view(self, checked: bool):
        """Toggle compact view mode"""
        self.config.set("compact_view", checked)
        self._update_compact_view()

    def _update_compact_view(self):
        """Update toolbar icons based on compact view setting - hide buttons completely"""
        compact = self.config.get("compact_view", False)

        # Get the toolbar
        toolbar = getattr(self, 'main_toolbar', None)
        if not toolbar:
            # Fallback: find toolbar
            for widget in self.findChildren(QToolBar):
                if hasattr(self, 'new_record_action') and self.new_record_action in widget.actions():
                    toolbar = widget
                    self.main_toolbar = toolbar  # Store it
                    break

        if not toolbar:
            return

        # Get list of actions to hide
        actions_to_hide = []
        if hasattr(self, 'new_record_action'):
            actions_to_hide.append(self.new_record_action)
        if hasattr(self, 'delete_record_action'):
            actions_to_hide.append(self.delete_record_action)
        if hasattr(self, 'add_field_action'):
            actions_to_hide.append(self.add_field_action)
        if hasattr(self, 'delete_field_action'):
            actions_to_hide.append(self.delete_field_action)
        if hasattr(self, 'bulk_add_action'):
            actions_to_hide.append(self.bulk_add_action)
        if hasattr(self, 'duplicate_record_action'):
            actions_to_hide.append(self.duplicate_record_action)
        if hasattr(self, 'join_action'):
            actions_to_hide.append(self.join_action)
        if hasattr(self, 'lock_form_action'):
            actions_to_hide.append(self.lock_form_action)

        # Hide/show actions directly - QAction has setVisible method
        for action in actions_to_hide:
            action.setVisible(not compact)

        # Also hide/show the widgets for visual consistency
        from PySide6.QtWidgets import QToolButton
        for action in actions_to_hide:
            widget = toolbar.widgetForAction(action)
            if widget:
                widget.setVisible(not compact)
            else:
                # Fallback: find button by iterating
                for button in toolbar.findChildren(QToolButton):
                    if button.defaultAction() == action:
                        button.setVisible(not compact)
                        break

        # Restore icons when not in compact view
        if not compact:
            if hasattr(self, 'new_record_action'):
                self.new_record_action.setIcon(QIcon(str(asset_path("add_row.png"))))
            if hasattr(self, 'delete_record_action'):
                self.delete_record_action.setIcon(QIcon(str(asset_path("delete_row.png"))))
            if hasattr(self, 'add_field_action'):
                self.add_field_action.setIcon(QIcon(str(asset_path("add_field.png"))))
            if hasattr(self, 'delete_field_action'):
                self.delete_field_action.setIcon(QIcon(str(asset_path("delete_field.png"))))
            if hasattr(self, 'lock_form_action'):
                # Restore icon based on state
                if self.lock_form_action.isChecked():
                    self.lock_form_action.setIcon(QIcon(str(asset_path("lock.png"))))
                else:
                    self.lock_form_action.setIcon(QIcon(str(asset_path("unlock.png"))))

        # Hide separators that are adjacent to hidden buttons
        # Get list of actions to hide for separator logic
        actions_to_hide = []
        if hasattr(self, 'new_record_action'):
            actions_to_hide.append(self.new_record_action)
        if hasattr(self, 'delete_record_action'):
            actions_to_hide.append(self.delete_record_action)
        if hasattr(self, 'add_field_action'):
            actions_to_hide.append(self.add_field_action)
        if hasattr(self, 'delete_field_action'):
            actions_to_hide.append(self.delete_field_action)
        if hasattr(self, 'bulk_add_action'):
            actions_to_hide.append(self.bulk_add_action)
        if hasattr(self, 'duplicate_record_action'):
            actions_to_hide.append(self.duplicate_record_action)
        if hasattr(self, 'join_action'):
            actions_to_hide.append(self.join_action)
        if hasattr(self, 'lock_form_action'):
            actions_to_hide.append(self.lock_form_action)

        all_actions = toolbar.actions()
        for i, action in enumerate(all_actions):
            if action.isSeparator():
                # Check if adjacent actions are hidden
                prev_hidden = False
                next_hidden = False

                if i > 0:
                    prev_action = all_actions[i - 1]
                    prev_hidden = prev_action in actions_to_hide and compact

                if i < len(all_actions) - 1:
                    next_action = all_actions[i + 1]
                    next_hidden = next_action in actions_to_hide and compact

                # Hide separator if adjacent to hidden actions
                if compact and (prev_hidden or next_hidden):
                    widget = toolbar.widgetForAction(action)
                    if widget:
                        widget.setVisible(False)
                elif not compact:
                    # Show separator when not in compact view
                    widget = toolbar.widgetForAction(action)
                    if widget:
                        widget.setVisible(True)

    def _toggle_collection_panel(self, checked: bool):
        """Toggle collection panel visibility"""
        self.config.set("visible_collection_panel", checked)
        self._apply_view_settings()

    def _toggle_show_key(self, checked: bool):
        """Toggle primary key column visibility"""
        self.config.set("show_key_column", checked)
        if hasattr(self, 'table_view') and self.table_view:
            self.table_view.setColumnHidden(0, not checked)

    def _toggle_expanded_view(self, checked: bool):
        """Toggle expanded view mode"""
        self.config.set("expanded_view", checked)
        # Apply table view settings to update view mode
        self._apply_table_view_settings()
        # Refresh the table if a collection is open
        if self.current_store:
            fields = self.current_store.list_fields()
            self.table_view.set_collection(self.current_store, fields)

    def _check_toolbar_overflow(self):
        """Check if toolbar has overflow and update overflow menu"""
        if not hasattr(self, 'overflow_button') or not hasattr(self, 'main_toolbar'):
            return

        toolbar = self.main_toolbar
        if not toolbar or not toolbar.isVisible():
            return

        # Wait for layout to update
        from PySide6.QtCore import QTimer
        QTimer.singleShot(10, lambda: self._do_check_overflow())

    def _do_check_overflow(self):
        """Actually perform overflow check after layout update"""
        if not hasattr(self, 'overflow_button') or not hasattr(self, 'main_toolbar'):
            return

        toolbar = self.main_toolbar
        if not toolbar:
            return

        # Get toolbar's visible width
        toolbar_width = toolbar.width()
        if toolbar_width == 0:
            return

        # Get all actions (excluding separators and overflow button)
        all_actions = toolbar.actions()
        hidden_actions = []

        # Find the overflow button widget to exclude it
        overflow_widget = None
        if hasattr(self, 'overflow_button'):
            overflow_widget = self.overflow_button

        # Check each action's widget position
        from PySide6.QtWidgets import QToolButton

        for action in all_actions:
            if action.isSeparator():
                continue

            widget = toolbar.widgetForAction(action)
            if not widget:
                # Try to find button widget
                for btn in toolbar.findChildren(QToolButton):
                    if btn.defaultAction() == action:
                        widget = btn
                        break

            if widget and widget != overflow_widget:
                # Reserve space for overflow button (40px to be safe)
                overflow_button_width = 40
                visible_threshold = toolbar_width - overflow_button_width

                # Check if widget is actually visible and if it extends beyond threshold
                # A widget is hidden if its right edge is beyond the visible threshold
                # OR if it's not visible at all (Qt might have hidden it)
                if not widget.isVisible():
                    # Widget is hidden by Qt's overflow mechanism
                    if action.text() or action.toolTip():
                        hidden_actions.append(action)
                else:
                    # Get widget's position relative to toolbar
                    widget_rect = widget.geometry()
                    if widget_rect.right() > visible_threshold:
                        # Widget extends beyond visible area
                        if action.text() or action.toolTip():
                            hidden_actions.append(action)

        # Update overflow menu
        self.overflow_menu.clear()

        if hidden_actions:
            # Add hidden actions to overflow menu
            for action in hidden_actions:
                if not action.isSeparator():
                    text = action.text() or action.toolTip() or "Action"
                    menu_action = self.overflow_menu.addAction(action.icon(), text)
                    menu_action.triggered.connect(action.trigger)
                    menu_action.setEnabled(action.isEnabled())
            # Show overflow button
            self.overflow_button.setVisible(True)
            self.overflow_button.setToolTip(f"More options ({len(hidden_actions)} hidden)")
        else:
            # Hide overflow button
            self.overflow_button.setVisible(False)

    def _apply_view_settings(self):
        """Apply view settings from config"""
        # Apply collection panel visibility
        visible = self.config.get("visible_collection_panel", True)
        if hasattr(self, 'sidebar_widget'):
            self.sidebar_widget.setVisible(visible)

        # Apply key column visibility
        show_key = self.config.get("show_key_column", True)
        if hasattr(self, 'table_view') and self.table_view:
            self.table_view.setColumnHidden(0, not show_key)
        if hasattr(self, 'show_key_action'):
            self.show_key_action.setChecked(show_key)

        # Update compact view
        if hasattr(self, 'compact_view_action'):
            self.compact_view_action.setChecked(self.config.get("compact_view", False))
        self._update_compact_view()

        # Update expanded view
        if hasattr(self, 'expanded_view_action'):
            self.expanded_view_action.setChecked(self.config.get("expanded_view", False))

        # Apply table view settings
        self._apply_table_view_settings()

    def _apply_table_view_settings(self):
        """Apply table view settings from config"""
        if not hasattr(self, 'table_view') or not self.table_view:
            return

        # Check if expanded view is enabled
        expanded_view = self.config.get("expanded_view", False)

        if expanded_view:
            # Expanded view: maximize everything to show all data
            self.table_view.setWordWrap(True)  # Enable word wrap for text fields
            # Set row height to auto-resize based on content
            self.table_view.verticalHeader().setSectionResizeMode(
                self.table_view.verticalHeader().ResizeMode.ResizeToContents
            )
            # Set column width to auto-resize based on content
            self.table_view.horizontalHeader().setSectionResizeMode(
                self.table_view.horizontalHeader().ResizeMode.ResizeToContents
            )
            # Resize columns to contents
            self.table_view.resizeColumnsToContents()
            # Set minimum row height to accommodate wrapped text and checkboxes
            # Need at least 24px to fit checkbox (20px) + 2px padding on each side
            self.table_view.verticalHeader().setMinimumSectionSize(24)
        else:
            # Normal view: apply configured settings
            self.table_view.setWordWrap(False)  # Disable word wrap
            # Set row height
            row_height = self.config.get("table_row_height", 24)
            self.table_view.verticalHeader().setDefaultSectionSize(row_height)
            self.table_view.verticalHeader().setSectionResizeMode(
                self.table_view.verticalHeader().ResizeMode.Fixed
            )
            # Apply to all existing rows
            for row in range(self.table_view.model.rowCount()):
                self.table_view.setRowHeight(row, row_height)

            # Apply default column width (for new columns)
            col_width = self.config.get("column_width_default", 120)
            # Set default width for horizontal header
            self.table_view.horizontalHeader().setDefaultSectionSize(col_width)
            self.table_view.horizontalHeader().setSectionResizeMode(
                self.table_view.horizontalHeader().ResizeMode.Interactive
            )
            # Apply to columns that are smaller than default (but respect auto-sized larger columns)
            for col in range(self.table_view.model.columnCount()):
                current_width = self.table_view.columnWidth(col)
                if col == 0:
                    # Primary key column - keep at 60
                    continue
                if current_width < col_width:
                    self.table_view.setColumnWidth(col, col_width)

        # Apply font size (always applies)
        font_size = self.config.get("font_size", 10)
        font = self.table_view.font()
        font.setPointSize(font_size)
        self.table_view.setFont(font)

        # Ensure vertical header width is maintained (row numbers visibility)
        self.table_view.verticalHeader().setFixedWidth(70)  # Ensure row numbers are visible

        # Also apply to form view
        if hasattr(self, 'form_view') and self.form_view:
            form_font = self.form_view.font()
            form_font.setPointSize(font_size)
            self.form_view.setFont(form_font)

    def _show_shortcuts(self):
        """Show keyboard shortcuts dialog"""
        from src.ui.shortcuts_dialog import ShortcutsDialog
        dialog = ShortcutsDialog(self)
        dialog.exec()

    def _manual_check_for_updates(self):
        """Manually check for updates from Tools menu"""
        from PySide6.QtWidgets import QMessageBox

        # Show non-blocking "checking" indicator with Cancel so the user can dismiss it
        checking_msg = QMessageBox(self)
        checking_msg.setWindowTitle("Checking for Updates")
        checking_msg.setText("Checking for updates...")
        checking_msg.setStandardButtons(QMessageBox.StandardButton.Cancel)
        checking_msg.show()

        thread = UpdateCheckWorker(self)
        self.update_check_threads.append(thread)

        def _dismiss():
            checking_msg.close()

        def on_update_available(update_info):
            _dismiss()
            self._show_update_dialog(update_info)

        def on_no_update():
            _dismiss()
            QMessageBox.information(
                self,
                "No Updates",
                f"You are running the latest version (v{VERSION})."
            )

        def on_error(error_msg):
            _dismiss()
            QMessageBox.warning(
                self,
                "Update Check Failed",
                f"Could not check for updates:\n{error_msg}"
            )

        def on_cancel(_button):
            """Stop the check thread when the user clicks Cancel"""
            if thread.isRunning():
                thread.terminate()
                thread.wait(1000)
            self._cleanup_thread(thread)

        checking_msg.buttonClicked.connect(on_cancel)
        thread.update_available.connect(on_update_available)
        thread.no_update.connect(on_no_update)
        thread.error.connect(on_error)
        # Always dismiss the dialog and clean up, even if no signal was emitted
        thread.finished.connect(_dismiss)
        thread.finished.connect(lambda: self._cleanup_thread(thread))
        thread.start()

    def _check_for_updates_async(self):
        """Check for updates asynchronously on startup (if enabled)"""
        thread = UpdateCheckWorker(self)
        self.update_check_threads.append(thread)

        def on_update_available(update_info):
            ignored_versions = self.config.get("update_ignored_versions", [])
            if update_info['version'] not in ignored_versions:
                self._show_update_dialog(update_info)

        # On startup: silently ignore "no update" and errors — don't interrupt the user
        thread.update_available.connect(on_update_available)
        thread.no_update.connect(lambda: None)
        thread.error.connect(lambda _: None)
        thread.finished.connect(lambda: self._cleanup_thread(thread))
        thread.start()

    def _cleanup_thread(self, thread):
        """Clean up a finished update check thread"""
        if thread in self.update_check_threads:
            self.update_check_threads.remove(thread)
        thread.deleteLater()

    def _show_update_dialog(self, update_info: dict):
        """Show update dialog and handle user response"""
        dialog = UpdateDialog(update_info, self)
        if dialog.exec():
            # User chose to download - start automatic download and installation
            download_url = update_info.get('download_url')
            if download_url:
                # Show progress dialog and download/install automatically
                progress_dialog = UpdateProgressDialog(download_url, self)
                progress_dialog.exec()
            else:
                # Fallback to manual download if no direct URL
                self._open_download_url(update_info)
        elif dialog.ignored:
            # User chose to ignore this version
            ignored_versions = self.config.get("update_ignored_versions", [])
            if update_info['version'] not in ignored_versions:
                ignored_versions.append(update_info['version'])
                self.config.set("update_ignored_versions", ignored_versions)

    def _open_download_url(self, update_info: dict):
        """Open the download URL in the default browser (fallback method)"""
        import webbrowser
        download_url = update_info.get('download_url')
        release_url = update_info.get('url')

        # Prefer direct download URL, fallback to release page
        url = download_url if download_url else release_url
        if url:
            webbrowser.open(url)
            QMessageBox.information(
                self,
                "Download Started",
                "The update download page has been opened in your browser.\n\n"
                "After downloading, close Quartz and run the installer to update."
            )
        else:
            QMessageBox.warning(
                self,
                "Download Error",
                "Could not find download URL for this update."
            )

    def _show_settings(self):
        """Show settings/preferences window"""
        from src.ui.preferences_dialog import PreferencesDialog

        dialog = PreferencesDialog(self, self.config)
        if dialog.exec():
            # Reapply theme if it changed
            self._apply_theme()
            # Apply table view settings
            self._apply_table_view_settings()
            # Refresh views to apply appearance changes
            if self.current_store:
                fields = self.current_store.list_fields()
                self.table_view.set_collection(self.current_store, fields)
                self.form_view.set_collection(self.current_store, fields)

    def _refresh_all(self):
        """Refresh everything - deselect collections and reload"""
        # Deselect current collection
        self._deselect_collection()

        # Reload collections list
        self._load_collections()

        # Clear any search/filter state
        if hasattr(self, 'search_box'):
            self.search_box.clear()

        # Clear collection filters and sorting
        if hasattr(self, 'collection_filters'):
            self.collection_filters.clear()
        if hasattr(self, 'collection_sorting'):
            self.collection_sorting.clear()

        # Update status
        self.statusBar().showMessage("Refreshed - all collections reloaded", 3000)

    def _confirm_destructive_action(self, title: str, message: str) -> bool:
        """Show confirmation dialog requiring user to type 'delete'"""
        from PySide6.QtWidgets import (
            QDialog,
            QHBoxLayout,
            QLabel,
            QLineEdit,
            QPushButton,
            QVBoxLayout,
        )

        dialog = QDialog(self)
        dialog.setWindowTitle(title)
        dialog.setMinimumWidth(400)

        layout = QVBoxLayout(dialog)

        # Warning message
        warning_label = QLabel(message)
        warning_label.setWordWrap(True)
        layout.addWidget(warning_label)

        # Instruction
        instruction_label = QLabel("Type 'delete' to confirm:")
        layout.addWidget(instruction_label)

        # Input field
        confirm_input = QLineEdit()
        confirm_input.setPlaceholderText("Type 'delete' here...")
        layout.addWidget(confirm_input)

        # Buttons
        button_layout = QHBoxLayout()
        button_layout.addStretch()

        cancel_btn = QPushButton("Cancel")
        cancel_btn.setProperty("class", "secondary")
        cancel_btn.clicked.connect(dialog.reject)

        confirm_btn = QPushButton("Confirm")
        confirm_btn.setDefault(True)
        confirm_btn.setEnabled(False)  # Disabled until 'delete' is typed

        # Enable confirm button only when 'delete' is typed
        def on_text_changed(text):
            confirm_btn.setEnabled(text.strip().lower() == "delete")

        confirm_input.textChanged.connect(on_text_changed)
        confirm_btn.clicked.connect(dialog.accept)

        button_layout.addWidget(cancel_btn)
        button_layout.addWidget(confirm_btn)
        layout.addLayout(button_layout)

        # Apply theme
        if hasattr(self, 'styleSheet'):
            dialog.setStyleSheet(self.styleSheet())

        return dialog.exec() == QDialog.Accepted

    def _delete_all_records(self):
        """Delete all records from the current collection"""
        if not self.current_store or not self.current_collection:
            QMessageBox.information(self, "Info", "Please select a collection first")
            return

        # Get record count
        record_count = self.current_store.count_records()
        if record_count == 0:
            QMessageBox.information(self, "Info", "No records to delete")
            return

        # Confirm with typing 'delete'
        confirmed = self._confirm_destructive_action(
            "Delete All Records",
            f"⚠️ WARNING: This will permanently delete ALL {record_count} record(s) from '{self.current_collection}'.\n\n"
            f"This action cannot be undone.\n\n"
            f"Are you sure you want to proceed?"
        )

        if not confirmed:
            return

        try:
            # Delete all records using SQL
            self.current_store.connect()
            cursor = self.current_store.conn.cursor()
            cursor.execute("DELETE FROM records")

            # Update FTS index
            self.current_store.update_fts_index()

            self.current_store.conn.commit()

            # Refresh views
            if hasattr(self, 'table_view') and self.table_view:
                self.table_view.model._refresh_data()
            if hasattr(self, 'form_view') and self.form_view:
                self.form_view.new_record()
            self._update_navigation()

            QMessageBox.information(
                self, "Success",
                f"Deleted all {record_count} record(s) from '{self.current_collection}'."
            )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to delete all records:\n{str(e)}")

    def _delete_all_collections(self):
        """Delete all collections"""
        # Get collection count
        collections = self.workspace.list_collections()
        collection_count = len(collections)

        if collection_count == 0:
            QMessageBox.information(self, "Info", "No collections to delete")
            return

        # Confirm with typing 'delete'
        confirmed = self._confirm_destructive_action(
            "Delete All Collections",
            f"⚠️ WARNING: This will permanently delete ALL {collection_count} collection(s).\n\n"
            f"This action cannot be undone.\n\n"
            f"All data, fields, and records will be lost.\n\n"
            f"Are you sure you want to proceed?"
        )

        if not confirmed:
            return

        # Confirm again with a simple yes/no
        reply = QMessageBox.question(
            self,
            "Final Confirmation",
            f"You are about to delete {collection_count} collection(s).\n\n"
            f"This is your last chance to cancel.\n\n"
            f"Proceed with deletion?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )

        if reply != QMessageBox.Yes:
            return

        try:
            deleted_count = 0
            errors = []

            # Close current collection if open
            if self.current_store:
                self.current_store.close()
                self.current_store = None
                self.current_collection = None

            # Delete each collection
            for collection_name in list(collections):
                try:
                    # Don't create backup for bulk delete
                    self.workspace.delete_collection(collection_name, backup=False)
                    deleted_count += 1
                except Exception as e:
                    errors.append(f"{collection_name}: {str(e)}")

            # Refresh UI
            self._load_collections()
            self._deselect_collection()

            # Show results
            if errors:
                QMessageBox.warning(
                    self, "Deletion Complete with Errors",
                    f"Deleted {deleted_count} of {collection_count} collection(s).\n\n"
                    f"Errors:\n" + "\n".join(errors[:5])
                )
            else:
                QMessageBox.information(
                    self, "Success",
                    f"Deleted all {deleted_count} collection(s)."
                )
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to delete all collections:\n{str(e)}")

    def _undo(self):
        """Undo the last action"""
        if not self.undo_history:
            return

        # Get the last command
        command = self.undo_history.pop()

        try:
            # Execute undo
            command.undo()

            # Move to redo history
            self.redo_history.append(command)
            if len(self.redo_history) > self.max_history:
                self.redo_history.pop(0)

            # Update button states
            self._update_undo_redo_buttons()

            # Refresh views
            if self.current_store:
                if hasattr(self, 'table_view') and self.table_view:
                    self.table_view.model._refresh_data()
                if hasattr(self, 'form_view') and self.form_view:
                    if hasattr(self.form_view, 'current_record_id') and self.form_view.current_record_id:
                        self.form_view.load_record(self.form_view.current_record_id)
                self._update_navigation()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to undo: {str(e)}")
            # Put command back if undo failed
            self.undo_history.append(command)

    def _redo(self):
        """Redo the last undone action"""
        if not self.redo_history:
            return

        # Get the last undone command
        command = self.redo_history.pop()

        try:
            # Execute redo
            command.redo()

            # Move back to undo history
            self.undo_history.append(command)
            if len(self.undo_history) > self.max_history:
                self.undo_history.pop(0)

            # Update button states
            self._update_undo_redo_buttons()

            # Refresh views
            if self.current_store:
                if hasattr(self, 'table_view') and self.table_view:
                    self.table_view.model._refresh_data()
                if hasattr(self, 'form_view') and self.form_view:
                    if hasattr(self.form_view, 'current_record_id') and self.form_view.current_record_id:
                        self.form_view.load_record(self.form_view.current_record_id)
                self._update_navigation()
        except Exception as e:
            QMessageBox.warning(self, "Error", f"Failed to redo: {str(e)}")
            # Put command back if redo failed
            self.redo_history.append(command)

    def _update_undo_redo_buttons(self):
        """Update undo/redo button enabled states"""
        if hasattr(self, 'undo_action'):
            self.undo_action.setEnabled(len(self.undo_history) > 0)
        if hasattr(self, 'redo_action'):
            self.redo_action.setEnabled(len(self.redo_history) > 0)

    def _add_to_history(self, command):
        """Add a command to undo history"""
        self.undo_history.append(command)
        if len(self.undo_history) > self.max_history:
            self.undo_history.pop(0)
        # Clear redo history when new action is performed
        self.redo_history.clear()
        self._update_undo_redo_buttons()

    def _show_audit_trail(self):
        """Show the Audit Trail dialog"""
        from src.ui.audit_trail_dialog import AuditTrailDialog

        if not self.undo_history and not self.redo_history:
            QMessageBox.information(
                self,
                "Audit Trail",
                "No changes have been recorded yet.\n\n"
                "Make edits to records and the audit trail will appear here.",
            )
            return

        dlg = AuditTrailDialog(self.undo_history, self.redo_history, parent=self)
        result = dlg.exec()

        # result > 0 means "Undo to here" was clicked with that many steps
        if result > 0:
            for _ in range(result):
                self._undo()

    def _prev_record(self):
        """Navigate to previous record"""
        if not self.current_store:
            return

        # Get current selection in table
        selection = self.table_view.selectionModel().selectedRows()
        if selection:
            current_row = selection[0].row()
            model = self.table_view.model
            prev_row = current_row - 1
            if 0 <= prev_row < model.rowCount():
                self.table_view.selectRow(prev_row)
                self.table_view.scrollTo(model.index(prev_row, 0))
                # If in form view, load the record
                if self.content_stack.currentIndex() == 2:  # Form view
                    record = model._get_record(prev_row)
                    if record:
                        self.form_view.load_record(record["id"])

    def _next_record(self):
        """Navigate to next record"""
        if not self.current_store:
            return

        # Get current selection in table
        selection = self.table_view.selectionModel().selectedRows()
        if selection:
            current_row = selection[0].row()
            model = self.table_view.model
            next_row = current_row + 1
            if next_row < model.rowCount():
                # Select next row
                self.table_view.selectRow(next_row)
                self.table_view.scrollTo(model.index(next_row, 0))
                # If in form view, load the record
                if self.content_stack.currentIndex() == 2:  # Form view
                    record = model._get_record(next_row)
                    if record:
                        self.form_view.load_record(record["id"])

    def _update_navigation(self):
        """Update navigation label"""
        if not self.current_store:
            self.nav_label.setText("No collection")
            return

        model = self.table_view.model
        # Use _total_count for virtualized collections (records list is empty in that mode)
        total = model._total_count if (model._virtualized and model._total_count > 500) else len(model.records)
        filtered = len(model.filtered_records)

        if model._is_filtered:
            self.nav_label.setText(f"Records: {filtered} of {total} (filtered)")
        else:
            self.nav_label.setText(f"Records: {total}")

    def _on_record_saved(self, record_id):
        """Handle record saved from form view"""
        # Refresh table view to show new/updated record
        if self.table_view.model and self.table_view.model.store:
            self.table_view.model._refresh_data()
            # Force table view to update its display
            self.table_view.viewport().update()
        # Update navigation counter
        self._update_navigation()

        # If in table view, select the newly saved record
        if self.table_toggle.isChecked():
            model = self.table_view.model
            if model and model.filtered_records:
                for i, record in enumerate(model.filtered_records):
                    if record["id"] == record_id:
                        self.table_view.selectRow(i)
                        # Scroll to the selected row
                        self.table_view.scrollTo(model.index(i, 0))
                        break

    def eventFilter(self, obj, event):
        """Event filter to detect clicks on empty space in top bar and collections list"""
        from PySide6.QtGui import QMouseEvent

        # Handle mouse press on top bar widget (check if it exists first)
        if hasattr(self, 'top_bar_widget') and obj == self.top_bar_widget and event.type() == QEvent.MouseButtonPress:
            if isinstance(event, QMouseEvent) and event.button() == Qt.LeftButton:
                # Check if click is on empty space (not on child widgets)
                child_at_pos = obj.childAt(event.pos())
                if child_at_pos == obj or child_at_pos is None:
                    # Clicked on empty space, deselect collection
                    self._deselect_collection()
                    # Also clear selection in collections list
                    if hasattr(self, 'collections_list'):
                        self.collections_list.clearSelection()
                    return True

        # Handle mouse press on collections list (for clicks on empty space and right-clicks)
        if hasattr(self, 'collections_list') and obj == self.collections_list and event.type() == QEvent.MouseButtonPress:
            if isinstance(event, QMouseEvent) and event.button() == Qt.RightButton:
                # Store current selection BEFORE Qt processes the right-click
                current_item = self.collections_list.currentItem()
                if current_item:
                    self._right_click_selected_row = self.collections_list.currentRow()
                else:
                    self._right_click_selected_row = -1

                # Block signals IMMEDIATELY to prevent selection change
                self.collections_list.blockSignals(True)
                self._right_click_in_progress = True

                # Restore selection multiple times to ensure it sticks
                # First restore immediately (before Qt processes)
                if self._right_click_selected_row >= 0 and self._right_click_selected_row < self.collections_list.count():
                    self.collections_list.setCurrentRow(self._right_click_selected_row)

                # Then restore again after Qt processes the event
                from PySide6.QtCore import QTimer
                def restore_selection():
                    if hasattr(self, '_right_click_selected_row') and self._right_click_selected_row >= 0:
                        if self._right_click_selected_row < self.collections_list.count():
                            self.collections_list.setCurrentRow(self._right_click_selected_row)
                    self.collections_list.blockSignals(False)
                    self._right_click_in_progress = False

                # Use multiple timers to ensure restoration happens
                QTimer.singleShot(0, restore_selection)
                QTimer.singleShot(10, restore_selection)  # Backup restoration
            elif isinstance(event, QMouseEvent) and event.button() == Qt.LeftButton:
                # Check if click is on an item
                item = self.collections_list.itemAt(event.pos())
                if item is None:
                    # Clicked on empty space, clear selection
                    self.collections_list.clearSelection()
                    # This will trigger itemSelectionChanged which calls _deselect_collection
                    return True

        return super().eventFilter(obj, event)

    def resizeEvent(self, event):
        """Handle window resize - check toolbar overflow"""
        super().resizeEvent(event)
        # Check toolbar overflow after resize
        from PySide6.QtCore import QTimer
        QTimer.singleShot(50, self._check_toolbar_overflow)

    def showEvent(self, event):
        """Handle window show - check toolbar overflow"""
        super().showEvent(event)
        # Check overflow after window is shown
        from PySide6.QtCore import QTimer
        QTimer.singleShot(300, self._check_toolbar_overflow)

    def closeEvent(self, event):
        """Handle window close"""
        # Wait for any running update check threads to finish
        for thread in self.update_check_threads[:]:  # Copy list to avoid modification during iteration
            if thread.isRunning():
                thread.wait(1000)  # Wait up to 1 second for thread to finish
            if thread in self.update_check_threads:
                self.update_check_threads.remove(thread)
            thread.deleteLater()

        if self.current_store:
            self.current_store.close()
        event.accept()
